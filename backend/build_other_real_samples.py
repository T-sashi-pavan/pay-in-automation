import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import random

NON_SLAB_HEADERS = [
    "LOB", "File Type", "Insurance Company", "Product", "Policy Type", "Plan Type", 
    "Sub Product", "Class", "Sub Class", "Make", "Model", "Fuel Type", "CPA Status", 
    "NCB Status", "Vehicle Age From", "Vehicle Age To", "Partner Type", "Source", 
    "Zone", "State", "RTO", "Effective Date", "Remarks", "Payin OD", "Payin TP", 
    "Payin Net", "Payout OD", "Payout TP", "Payout Net", "Payin Reward", 
    "Payout Reward", "Payin Scheme", "Payout Scheme"
]

SLAB_HEADERS = [
    "LOB", "File Type", "Insurance Company", "Product", "Policy Type", "Plan Type", 
    "Sub Product", "Class", "Sub Class", "Make", "Model", "Fuel Type", "CPA Status", 
    "NCB Status", "Vehicle Age From", "Vehicle Age To", "Partner Type", "Source", 
    "Zone", "State", "RTO", "Effective Date", "Remarks", "Payin Type", "Premium Type", 
    "Payin Slab From", "Payin Slab To", "Payin OD", "Payin TP", "Payin Net", 
    "Payout OD", "Payout TP", "Payout Net"
]

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
DATA_FONT = Font(size=10, name="Segoe UI")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

def autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 60)

def split_slabs_by_premium_type(slabs):
    new_slabs = []
    for slab in slabs:
        od = slab.get("Payin OD") or 0.0
        tp = slab.get("Payin TP") or 0.0
        net = slab.get("Payin Net") or 0.0
        
        # Determine which fields are positive
        positive_types = []
        if od > 0.0:
            positive_types.append("OD")
        if tp > 0.0:
            positive_types.append("TP")
        if net > 0.0:
            positive_types.append("NET")
            
        if len(positive_types) > 1:
            for t in positive_types:
                new_slabs.append({
                    "Payin Type": slab.get("Payin Type"),
                    "Premium Type": t,
                    "Payin Slab From": slab.get("Payin Slab From"),
                    "Payin Slab To": slab.get("Payin Slab To"),
                    "Payin OD": od if t == "OD" else 0.0,
                    "Payin TP": tp if t == "TP" else 0.0,
                    "Payin Net": net if t == "NET" else 0.0
                })
        else:
            t = positive_types[0] if positive_types else (slab.get("Premium Type") or "NET")
            new_slabs.append({
                "Payin Type": slab.get("Payin Type"),
                "Premium Type": t,
                "Payin Slab From": slab.get("Payin Slab From"),
                "Payin Slab To": slab.get("Payin Slab To"),
                "Payin OD": od if t == "OD" else 0.0,
                "Payin TP": tp if t == "TP" else 0.0,
                "Payin Net": net if t == "NET" else 0.0
            })
    return new_slabs

STATE_ZONE_RTO_MAP = {
    "AP": {"zone": "Zone B", "rto": "AP-02"},
    "TN": {"zone": "Zone A", "rto": "TN-01"},
    "KA": {"zone": "Zone A", "rto": "KA-01"},
    "MH": {"zone": "Zone A", "rto": "MH-12"},
    "GJ": {"zone": "Zone B", "rto": "GJ-01"},
    "JH": {"zone": "Zone C", "rto": "JH-01"},
    "BR": {"zone": "Zone C", "rto": "BR-01"},
    "WB": {"zone": "Zone B", "rto": "WB-01"},
    "DL": {"zone": "Zone A", "rto": "DL-01"},
    "PB": {"zone": "Zone B", "rto": "PB-02"},
    "AS": {"zone": "Zone C", "rto": "AS-01"},
    "TR": {"zone": "Zone C", "rto": "TR-01"},
    "AR": {"zone": "Zone C", "rto": "AR-01"},
    "ML": {"zone": "Zone C", "rto": "ML-05"},
    "MZ": {"zone": "Zone C", "rto": "MZ-01"},
    "AN": {"zone": "Zone C", "rto": "AN-01"},
    "OR": {"zone": "Zone C", "rto": "OD-02"},
    "ALL": {"zone": "Zone A", "rto": "DL-01"}
}

PARTNER_TYPES_POOL = ["POSP", "MISP", "Broker", "OEM Dealer", "Direct Retail"]
SOURCES_POOL = ["online", "offline"]

def fill_with_realistic_value(h, val, row_dict):
    if h == "Vehicle Age From":
        to_val = row_dict.get("Vehicle Age To")
        if (val == 1 or val == "1") and (to_val == 1 or to_val == "1"):
            return 0

    if val not in [None, "", "ALL", "ALL partners", "NA", "all"]:
        return val
        
    state = row_dict.get("State") or row_dict.get("state") or "ALL"
    state_first = state.split(",")[0].strip()
    geo = STATE_ZONE_RTO_MAP.get(state_first, STATE_ZONE_RTO_MAP["ALL"])
    
    product = str(row_dict.get("Product") or "").lower()
    
    if h == "Make":
        if "2w" in product or "tw" in product or "two wheeler" in product:
            return random.choice(["HERO", "BAJAJ", "RE", "HONDA"])
        elif "gcv" in product or "lcv" in product or "hcv" in product:
            return random.choice(["TATA", "MAHINDRA", "ASHOK LEYLAND"])
        return random.choice(["MARUTI", "HYUNDAI", "TATA", "HONDA"])
        
    elif h == "Model":
        if "2w" in product or "tw" in product or "two wheeler" in product:
            return random.choice(["ACTIVA", "SPLENDOR", "PULSAR", "CLASSIC 350"])
        elif "gcv" in product or "lcv" in product or "hcv" in product:
            return random.choice(["ACE", "BOLERO MAXI", "DOST"])
        return random.choice(["SWIFT", "ALTO 800", "NEXON", "BALENO"])
        
    elif h == "Fuel Type":
        if "2w" in product or "tw" in product or "two wheeler" in product:
            return "petrol"
        return random.choice(["petrol", "diesel", "CNG"])
        
    elif h == "CPA Status":
        policy_type = str(row_dict.get("Policy Type") or "").lower()
        if "third party" in policy_type or "tp" in policy_type:
            return "NO"
        return random.choice(["YES", "NO"])
        
    elif h == "NCB Status":
        file_type = str(row_dict.get("File Type") or "").lower()
        if "new" in file_type:
            return "NO"
        return random.choice(["YES", "NO"])
        
    elif h == "Partner Type":
        return random.choice(PARTNER_TYPES_POOL)
        
    elif h == "Source":
        return random.choice(SOURCES_POOL)
        
    elif h == "Zone":
        return geo["zone"]
        
    elif h == "RTO":
        return geo["rto"]
        
    elif h == "Vehicle Age From":
        return 1
        
    elif h == "State":
        return random.choice(["MH", "GJ", "AP", "KA", "TN", "DL"])
        
    return val

def generate_file(out_path, insurer_name, non_slab_rows, slab_rules, slab_sheet_name="SLAB"):
    random.seed(42)
    wb = openpyxl.Workbook()
    
    # 1. NON SLAB
    ws_ns = wb.active
    ws_ns.title = "NON SLAB"
    ws_ns.append(NON_SLAB_HEADERS)
    for cell in ws_ns[1]:
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN, border=THIN_BORDER)
        
    for r_idx, r in enumerate(non_slab_rows, start=2):
        row_values = []
        for h in NON_SLAB_HEADERS:
            val = r.get(h)
            if val is None:
                if h == "Payout OD" and r.get("Payin OD") is not None:
                    val = r.get("Payin OD") * 0.8
                elif h == "Payout TP" and r.get("Payin TP") is not None:
                    val = r.get("Payin TP") * 0.8
                elif h == "Payout Net" and r.get("Payin Net") is not None:
                    val = r.get("Payin Net") * 0.8
                elif h == "Payout Reward" and r.get("Payin Reward") is not None:
                    val = r.get("Payin Reward") * 0.8
                elif h == "Payout Scheme" and r.get("Payin Scheme") is not None:
                    val = r.get("Payin Scheme") * 0.8
                else:
                    val = "ALL" if h in ["Plan Type", "Make", "Model", "Fuel Type", "CPA Status", "NCB Status", "Source", "Zone", "RTO"] else None
            val = fill_with_realistic_value(h, val, r)
            row_values.append(val)
        ws_ns.append(row_values)
        for col_idx in range(1, len(NON_SLAB_HEADERS) + 1):
            cell = ws_ns.cell(row=r_idx, column=col_idx)
            style_cell(cell, alignment=CENTER_ALIGN, border=THIN_BORDER)
            
    # 2. SLAB
    ws_sl = wb.create_sheet(slab_sheet_name)
    ws_sl.append(SLAB_HEADERS)
    for cell in ws_sl[1]:
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN, border=THIN_BORDER)
        
    current_row = 2
    for r in slab_rules:
        slabs = split_slabs_by_premium_type(r["slabs"])
        num_slabs = len(slabs)
        
        # Write sub-rows
        for i, slab in enumerate(slabs):
            row_idx = current_row + i
            # Write business columns (LOB to Remarks)
            for col_idx, h in enumerate(SLAB_HEADERS[:23]):
                val = r.get(h)
                if val is None:
                    val = "ALL" if h in ["Plan Type", "Make", "Model", "Fuel Type", "CPA Status", "NCB Status", "Source", "Zone", "RTO"] else None
                val = fill_with_realistic_value(h, val, r)
                cell = ws_sl.cell(row=row_idx, column=col_idx + 1, value=val)
                style_cell(cell, alignment=CENTER_ALIGN, border=THIN_BORDER)
                
            # Write slab columns
            for col_idx, h in enumerate(SLAB_HEADERS[23:]):
                val = slab.get(h)
                if val is None:
                    if h == "Payout OD" and slab.get("Payin OD") is not None:
                        val = slab.get("Payin OD") * 0.8
                    elif h == "Payout TP" and slab.get("Payin TP") is not None:
                        val = slab.get("Payin TP") * 0.8
                    elif h == "Payout Net" and slab.get("Payin Net") is not None:
                        val = slab.get("Payin Net") * 0.8
                cell = ws_sl.cell(row=row_idx, column=24 + col_idx, value=val)
                style_cell(cell, alignment=CENTER_ALIGN, border=THIN_BORDER)
                
        # Apply Merges for columns 1 to 23
        if num_slabs > 1:
            for col_idx in range(1, 24):
                ws_sl.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row + num_slabs - 1, end_column=col_idx)
                for offset in range(num_slabs):
                    cell = ws_sl.cell(row=current_row + offset, column=col_idx)
                    style_cell(cell, alignment=CENTER_ALIGN, border=THIN_BORDER)
                    
        current_row += num_slabs
        
    autosize(ws_ns)
    autosize(ws_sl)
    try:
        wb.save(out_path)
        print(f"Generated {insurer_name} Excel file at: {out_path}")
    except PermissionError:
        print(f"ERROR: Permission denied when writing to {out_path}. Please close this file in Microsoft Excel and run again!")

def build_all():
    base_dir = r"c:\Desktop\ALGONOX\PAY-IN-AUTOMATION"
    
    # ---------------------------------------------
    # 1. CHOLAMANDALAM
    # ---------------------------------------------
    chola_ns = [
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "MISC-D TRACTOR", 
         "Policy Type": "Standalone Own Damage", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Agri/Tractors Only", "Payin OD": 12.5, "Payin TP": 12.5, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "MISC-D TRACTOR", 
         "Policy Type": "Standalone Own Damage", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "TN", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Agri/Tractors Only", "Payin OD": 10.0, "Payin TP": 12.5, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "Renewal", "Insurance Company": "Chola", "Product": "MISC-D TRACTOR", 
         "Policy Type": "Standalone Own Damage", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Agri/Tractors Only", "Payin OD": 15.0, "Payin TP": 12.5, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "1_GCCV_3W", "Make": "BAJAJ", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "KA", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "3-Wheeler Pack", "Payin OD": 7.5, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "GCCV", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "1_GCCV_3W", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "KA", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "3-Wheeler Act", "Payin OD": 0.0, "Payin TP": 32.5, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC Pack New", "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC Act New", "Payin OD": 0.0, "Payin TP": 10.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "Renewal", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 6, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "MH", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC Pack Renewal", "Payin OD": 22.5, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "Renewal", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 6, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "MH", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC Act Renewal", "Payin OD": 0.0, "Payin TP": 12.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "2_UPTO_3.5T", "Make": "TATA", "Model": "Ace", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "TN", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Ace Pack", "Payin OD": 45.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0}
    ]
    chola_sl = [
        # Rule 1: GCCV Slabs AP
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "Intra", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "GCCV GVW Slabs",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 3500, "Payin OD": 47.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 17.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 22.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 12001, "Payin Slab To": 16000, "Payin OD": 22.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 16001, "Payin Slab To": 20000, "Payin OD": 17.5, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 2: GCCV Slabs TN
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "Yodha", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "TN", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "GCCV GVW Slabs TN",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 3500, "Payin OD": 45.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 12001, "Payin Slab To": 16000, "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 16001, "Payin Slab To": 20000, "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 3: GCCV Slabs KA
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "Jeeto", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "KA", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "GCCV GVW Slabs KA",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 3500, "Payin OD": 45.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 12.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 12.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 12001, "Payin Slab To": 16000, "Payin OD": 12.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 16001, "Payin Slab To": 20000, "Payin OD": 12.5, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 4: Private Car Slabs AP Comp
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC Slabs AP",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1001, "Payin Slab To": 1500, "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 30.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 5: Private Car Slabs TN Comp
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "TN", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC Slabs TN",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1001, "Payin Slab To": 1500, "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 30.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 6: Private Car Slabs KA Comp
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "KA", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC Slabs KA",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1001, "Payin Slab To": 1500, "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 30.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 7: Private Car Slabs MH Comp
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "MH", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC Slabs MH",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 17.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1001, "Payin Slab To": 1500, "Payin OD": 22.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 27.5, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 8: Private Car Slabs AP TP
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC TP Slabs AP",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 0.0, "Payin TP": 10.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1001, "Payin Slab To": 1500, "Payin OD": 0.0, "Payin TP": 20.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 0.0, "Payin TP": 20.0, "Payin Net": 0.0}
         ]},
        # Rule 9: Private Car Slabs TN TP
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "TN", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC TP Slabs TN",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 0.0, "Payin TP": 10.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1001, "Payin Slab To": 1500, "Payin OD": 0.0, "Payin TP": 25.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 0.0, "Payin TP": 30.0, "Payin Net": 0.0}
         ]},
        # Rule 10: Private Car Slabs KA TP
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Chola", "Product": "PRIVATE CAR", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "KA", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC TP Slabs KA",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 0.0, "Payin TP": 19.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1001, "Payin Slab To": 1500, "Payin OD": 0.0, "Payin TP": 19.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 0.0, "Payin TP": 20.0, "Payin Net": 0.0}
         ]}
    ]
    
    generate_file(os.path.join(base_dir, "CRM_Ready_Chola_RealData.xlsx"), "Chola", chola_ns, chola_sl)
    
    # ---------------------------------------------
    # 2. GO DIGIT
    # ---------------------------------------------
    digit_ns = [
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Private Car", 
         "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 3 Year TP", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "JH", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit PC Pack JH", "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 25.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Private Car", 
         "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 3 Year TP", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "BR", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit PC Pack BR", "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 25.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Private Car", 
         "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 3 Year TP", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "WB", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit PC Pack WB", "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 25.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 5 Year TP", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "GJ", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit TW Pack", "Payin OD": 30.0, "Payin TP": 0.0, "Payin Net": 30.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 5 Year TP", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit TW Pack AP", "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 20.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "School Bus", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "TN", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit Staff Bus TN", "Payin OD": 42.5, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "School Bus", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "GJ", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit Staff Bus GJ", "Payin OD": 55.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "School Bus", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "MH", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit Staff Bus MH", "Payin OD": 55.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "School Bus", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "DL", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit Staff Bus DL", "Payin OD": 55.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "School Bus", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "PB", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit Staff Bus PB", "Payin OD": 55.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0}
    ]
    digit_sl = [
        # Rule 1: Two Wheeler SATP AN
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Two Wheeler", 
         "Policy Type": "Third Party", "Plan Type": "1 Year OD + 1 Year TP", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "Hero", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AN", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "2W CC Slabs SATP",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1, "Payin Slab To": 180, "Payin OD": 70.0, "Payin TP": 33.99, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 181, "Payin Slab To": 350, "Payin OD": 40.0, "Payin TP": 33.99, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 351, "Payin Slab To": 999999, "Payin OD": 20.0, "Payin TP": 21.5, "Payin Net": 0.0}
         ]},
        # Rule 2: Two Wheeler 1+5 AN
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 5 Year TP", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "Hero", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AN", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "2W CC Slabs 1+5",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 180, "Payin OD": 40.0, "Payin TP": 10.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 181, "Payin Slab To": 350, "Payin OD": 40.0, "Payin TP": 17.5, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 351, "Payin Slab To": 999999, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 3: PC SATP CNG
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Private Car", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "CNG", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "PB", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC Slabs SATP CNG",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 0.0, "Payin TP": 2.5, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1001, "Payin Slab To": 1500, "Payin OD": 0.0, "Payin TP": 10.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 0.0, "Payin TP": 5.0, "Payin Net": 0.0}
         ]},
        # Rule 4: PC SATP Petrol
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Private Car", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "PB", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC Slabs SATP Petrol",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1, "Payin Slab To": 1000, "Payin OD": 0.0, "Payin TP": 41.5, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1001, "Payin Slab To": 999999, "Payin OD": 0.0, "Payin TP": 44.5, "Payin Net": 0.0}
         ]},
        # Rule 5: PC SATP Diesel
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "Private Car", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "New Vehicle", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "PB", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "PC CC Slabs SATP Diesel",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1, "Payin Slab To": 1500, "Payin OD": 0.0, "Payin TP": 18.5, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1501, "Payin Slab To": 999999, "Payin OD": 0.0, "Payin TP": 40.5, "Payin Net": 0.0}
         ]},
        # Rule 6: GCCV Slabs JH Comp
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "JH", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit GCV GVW Slabs JH",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 1600, "Payin OD": 60.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1601, "Payin Slab To": 2500, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2501, "Payin Slab To": 3500, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 7: GCCV Slabs BR Comp
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "BR", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit GCV GVW Slabs BR",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 1600, "Payin OD": 60.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1601, "Payin Slab To": 2500, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2501, "Payin Slab To": 3500, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 8: GCCV Slabs WB Comp
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "WB", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit GCV GVW Slabs WB",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 1600, "Payin OD": 60.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1601, "Payin Slab To": 2500, "Payin OD": 75.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2501, "Payin Slab To": 3500, "Payin OD": 50.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 40.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 70.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 9: GCCV Slabs OR Comp
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "GCCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "OR", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit GCV GVW Slabs OR",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 1600, "Payin OD": 60.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1601, "Payin Slab To": 2500, "Payin OD": 80.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2501, "Payin Slab To": 3500, "Payin OD": 50.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 40.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 40.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 10: GCCV Slabs JH TP
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Digit", "Product": "GCCV", 
         "Policy Type": "Third Party", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "JH", "RTO": "ALL", "Effective Date": "2025-05-21", 
         "Remarks": "Digit GCV GVW TP Slabs JH",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1, "Payin Slab To": 1600, "Payin OD": 0.0, "Payin TP": 35.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 1601, "Payin Slab To": 2500, "Payin OD": 0.0, "Payin TP": 50.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 2501, "Payin Slab To": 3500, "Payin OD": 0.0, "Payin TP": 20.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 0.0, "Payin TP": 17.5, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "TP", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 0.0, "Payin TP": 12.5, "Payin Net": 0.0}
         ]}
    ]
    
    generate_file(os.path.join(base_dir, "CRM_Ready_GoDigit_RealData.xlsx"), "Digit", digit_ns, digit_sl)
    
    # ---------------------------------------------
    # 3. TATA
    # ---------------------------------------------
    tata_ns = [
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Private Car", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "MARUTI", "Model": "ALTO 800", "Fuel Type": "petrol", "CPA Status": "YES", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 5, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "MH", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata PC Grid Maruti", "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Private Car", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "MARUTI", "Model": "SWIFT", "Fuel Type": "petrol", "CPA Status": "YES", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 5, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "MH", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata PC Grid Maruti SWIFT", "Payin OD": 17.5, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "Renewal", "Insurance Company": "Tata", "Product": "Private Car", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 6, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "ALL", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata PC Grid Renewal", "Payin OD": 10.0, "Payin TP": 2.5, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "Renewal", "Insurance Company": "Tata", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 6, "Vehicle Age To": 15, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "ALL", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata TW Grid Renewal", "Payin OD": 15.0, "Payin TP": 5.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Private Car", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "ALL", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata PC Grid New", "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "ALL", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata TW Grid New", "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "GJ", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata TW Grid GJ", "Payin OD": 10.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata TW Grid AP", "Payin OD": 34.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "KA", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata TW Grid KA", "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0},
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Two Wheeler", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "OR", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Tata TW Grid OR", "Payin OD": 10.0, "Payin TP": 0.0, "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0}
    ]
    tata_sl = [
        # Rule 1: GCV GJ Slabs
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "GCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "GJ", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "GCV Slabs GJ",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 30.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 3500, "Payin OD": 31.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 32.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 33.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 2: GCV AP Slabs
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "GCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "GCV Slabs AP",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 10.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 3500, "Payin OD": 12.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 16.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 12.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 3: GCV KA Slabs
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "GCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "KA", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "GCV Slabs KA",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 12.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 3500, "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 16.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 17.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 4: GCV OR Slabs
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "GCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "OR", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "GCV Slabs OR",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 3500, "Payin OD": 28.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 31.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 32.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 5: GCV BR Slabs
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "GCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "BR", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "GCV Slabs BR",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 3500, "Payin OD": 18.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 19.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 6: PCV Volume Slabs
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "PCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "ALL", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "PCV Volume Slabs GWP",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 1, "Payin Slab To": 50000, "Payin OD": 10.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 50001, "Payin Slab To": 100000, "Payin OD": 12.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 100001, "Payin Slab To": 999999, "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 7: Private Car Volume Slabs
        {"LOB": "Motor", "File Type": "New", "Insurance Company": "Tata", "Product": "Private Car", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "ALL", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "Private Car Volume Slabs GWP",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 1, "Payin Slab To": 20000, "Payin OD": 12.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 20001, "Payin Slab To": 50000, "Payin OD": 14.5, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 50001, "Payin Slab To": 999999, "Payin OD": 17.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 8: GCV GJ Slabs 1-5 yrs
        {"LOB": "Motor", "File Type": "Used", "Insurance Company": "Tata", "Product": "GCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 5, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "GJ", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "GCV Slabs GJ 1-5 yrs",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 3500, "Payin OD": 26.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 27.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 28.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 9: GCV AP Slabs 1-5 yrs
        {"LOB": "Motor", "File Type": "Used", "Insurance Company": "Tata", "Product": "GCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 5, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "GCV Slabs AP 1-5 yrs",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 8.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 3500, "Payin OD": 10.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 14.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 10.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]},
        # Rule 10: GCV KA Slabs 1-5 yrs
        {"LOB": "Motor", "File Type": "Used", "Insurance Company": "Tata", "Product": "GCV", 
         "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL", 
         "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL", 
         "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 5, "Partner Type": "ALL partners", 
         "Source": "ALL", "Zone": "ALL", "State": "KA", "RTO": "ALL", "Effective Date": "2025-07-01", 
         "Remarks": "GCV Slabs KA 1-5 yrs",
         "slabs": [
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 10.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 3500, "Payin OD": 13.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 3501, "Payin Slab To": 7500, "Payin OD": 14.0, "Payin TP": 0.0, "Payin Net": 0.0},
             {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7501, "Payin Slab To": 12000, "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0}
         ]}
    ]
    
    # We call generate_file for Tata with slab_sheet_name="SLAB (incremental)" so it matches classify_rule_type's specific Tata checks.
    generate_file(os.path.join(base_dir, "CRM_Ready_Tata_RealData.xlsx"), "Tata", tata_ns, tata_sl, slab_sheet_name="SLAB (incremental)")

if __name__ == "__main__":
    build_all()
