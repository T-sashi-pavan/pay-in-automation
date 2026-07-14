import re
import openpyxl
import pandas as pd
from io import BytesIO
from typing import List, Dict, Any, Tuple, Optional
from datetime import date
from backend.app.services.mapping_engine.mapper import ColumnMapper
from backend.app.services.normalizer.normalizer import ValueNormalizer
from backend.app.services.validator.validator import RuleValidator

def split_product_range(product_str: str) -> Tuple[str, Optional[str]]:
    if not product_str:
        return "", None
    
    # We look for where the range starts.
    pattern = r'([><=]+|' \
              r'\b(?:upto|up\s+to|above|below|under|exceeding)\b|' \
              r'\b\d+(?:\.\d+)?\s*(?:cc|t|ton|tons|kg|hp|kw|\+|-|to)\b)'
    
    match = re.search(pattern, product_str, re.IGNORECASE)
    if match:
        start_idx = match.start()
        prod = product_str[:start_idx].strip()
        sub = product_str[start_idx:].strip()
        
        # Clean product
        prod = re.sub(r'[-\s]+$', '', prod).strip()
        
        return prod, sub
        
    return product_str, None


def extract_numeric_range(text: str) -> Tuple[Optional[float], Optional[float]]:
    if not text:
        return None, None
    text_clean = str(text).replace(",", "").strip()
    
    # 1. Plus matches: "100001+" -> (100001, None)
    match_plus = re.search(r'(\d+(?:\.\d+)?)\s*\+', text_clean)
    if match_plus:
        return float(match_plus.group(1)), None
        
    # 2. Ranges: "3.5-7.5" or "50001 to 100000" -> (50001, 100000)
    match_range = re.search(r'(\d+(?:\.\d+)?)\s*(?:-|to)\s*(\d+(?:\.\d+)?)', text_clean, re.IGNORECASE)
    if match_range:
        return float(match_range.group(1)), float(match_range.group(2))
        
    # 3. Double boundary: ">3.5<=7.5" -> (3.5, 7.5)
    numbers = re.findall(r'\d+(?:\.\d+)?', text_clean)
    if len(numbers) >= 2:
        return float(numbers[0]), float(numbers[1])
    elif len(numbers) == 1:
        num = float(numbers[0])
        if ">" in text_clean:
            return num, None
        elif "<" in text_clean or "upto" in text_clean.lower() or "up to" in text_clean.lower() or "under" in text_clean.lower():
            return 1.0, num
            
    return None, None


def classify_rule_type(sheet_name: str, headers: List[str], base_rule: Dict[str, Any], flat_rates: Dict[str, Any]) -> str:
    sheet_lower = sheet_name.lower()
    sheet_lower = re.sub(r"non[\s\-_]*slab", "", sheet_lower)
    headers_lower = [str(h).lower() for h in headers]
    
    # Check for explicit slab indicators (slab, si slab, premium slab, threshold)
    # Excludes "from", "to", "upto", and "range" as they match eligibility headers
    slab_keywords = ["slab", "si slab", "premium slab", "threshold"]
    if any(kw in sheet_lower for kw in slab_keywords) or any(any(kw in h for kw in slab_keywords) for h in headers_lower):
        return "SLAB"
        
    if flat_rates.get("slab_from") is not None or flat_rates.get("slab_to") is not None:
        return "SLAB"
        
    return "NON_SLAB"


def is_slab_rule(sheet_name: str, headers: List[str], base_rule: Dict[str, Any], flat_rates: Dict[str, Any]) -> bool:
    return classify_rule_type(sheet_name, headers, base_rule, flat_rates) == "SLAB"


def is_rule_effectively_empty(rule: Dict[str, Any]) -> bool:
    """
    True if a parsed rule carries no usable identifying data at all — no
    business-key fields, no rates, no slab tiers with a real rate. This is a
    real failure mode for OCR-based PDF extraction (a page's column layout
    can be misread badly enough that every cell lands on the wrong column and
    every value ends up null) — such rows should be dropped rather than
    stored as if they were real data.
    """
    # insurance_company is deliberately excluded — it gets a blanket filename/company
    # fallback applied to every row regardless of whether anything real was extracted,
    # so its presence alone doesn't indicate the row carries real data.
    identity_fields = ("product", "policy_type", "sub_class", "state", "make", "model", "remarks")
    if any(rule.get(f) for f in identity_fields):
        return False

    rate_fields = ("payin_od", "payin_tp", "payin_net", "payin_reward", "payin_scheme")
    if any(rule.get(f) is not None for f in rate_fields):
        return False

    for slab in rule.get("slabs") or []:
        if any(slab.get(f) is not None for f in ("payin_od", "payin_tp", "payin_net", "slab_from", "slab_to")):
            return False

    return True


def check_duplicate_slab_ranges(slabs: List[Dict[str, Any]]) -> List[str]:
    """
    Real slab tiers must each cover a distinct (slab_from, slab_to) range
    PER premium_type (OD/TP/NET) — two tiers sharing an identical range AND
    the same premium_type is a parsing failure (e.g. a boundary never got
    captured and both tiers fell back to the same value), not a legitimate
    business case. premium_type is part of the key because one discount tier
    legitimately produces two rows sharing a boundary — one OD, one TP — which
    is not a duplicate.

    A null/null boundary gets its own check: the display layer (rule_serializer's
    _serialize_slab) paints every unresolved (None, None) tier with the same
    1/500000 placeholder, so two or more null-boundary tiers in one rule LOOK
    like duplicates to anyone viewing them even though they're distinct "unknown"
    values underneath — flag that case too rather than only exact value matches.

    Returns human-readable warnings for each real issue found, to surface via
    the existing validation_status/warnings mechanism instead of silently
    accepting the invalid data.
    """
    warnings: List[str] = []
    seen: Dict[Tuple[Any, Any, Any], int] = {}
    null_boundary_count = 0
    for slab in slabs:
        if slab.get("slab_from") is None and slab.get("slab_to") is None:
            null_boundary_count += 1
            continue
        key = (slab.get("slab_from"), slab.get("slab_to"), slab.get("premium_type"))
        seen[key] = seen.get(key, 0) + 1

    if null_boundary_count > 1:
        warnings.append(
            f"{null_boundary_count} slab tier(s) have an unresolved boundary — they will display with an "
            f"identical placeholder range until the source data/extraction is fixed"
        )

    for (slab_from, slab_to, premium_type), count in seen.items():
        if count > 1:
            warnings.append(
                f"Duplicate slab range detected ({slab_from}–{slab_to}, {premium_type} appears {count} times) — possible extraction error"
            )
    return warnings


def parse_tiered_rate_text(text: Any) -> Optional[List[Tuple[Optional[int], Optional[int], float]]]:
    """
    Detects vehicle-age-tiered text packed into a single pivot cell, e.g.
    "Age 0: 60%/ 85% \nAge >=1: 85%" or "Upto 0-2 Yrs - 45%\n3+ Yrs - 50%".
    Returns a list of (age_from, age_to, pct) tuples when >=2 recognizable
    age-tier lines are found, else None (so the caller falls back to the
    existing single-value parsing path unchanged).

    Heuristic note: when a tier line itself contains more than one percentage
    (e.g. "Age 0: 60%/ 85%"), the FIRST percentage found is used as the
    representative rate for that tier — the source grids don't disambiguate
    which sub-value applies without more context, so this is a documented
    best-effort choice, not a guaranteed-correct business rule.
    """
    if not text or not isinstance(text, str):
        return None

    segments = [s.strip() for s in re.split(r'[\n\r]+', text) if s.strip()]
    if len(segments) < 2:
        return None

    tiers: List[Tuple[Optional[int], Optional[int], float]] = []
    for seg in segments:
        # Only treat this as an age-tiered cell if every line references age/years —
        # avoids false positives on unrelated multi-line text (e.g. maker-specific notes).
        if not re.search(r'\bage\b|\byrs?\b|\byear', seg, re.IGNORECASE):
            return None

        age_from: Optional[int] = None
        age_to: Optional[int] = None
        m_ge = re.search(r'>=\s*(\d+)', seg)
        m_range = re.search(r'(\d+)\s*(?:to|-)\s*(\d+)', seg, re.IGNORECASE)
        m_plus = re.search(r'(\d+)\s*\+', seg)
        m_single = re.search(r'\b(\d+)\b', seg)
        if m_ge:
            age_from = int(m_ge.group(1))
        elif m_range:
            age_from, age_to = int(m_range.group(1)), int(m_range.group(2))
        elif m_plus:
            age_from = int(m_plus.group(1))
        elif m_single:
            age_from = age_to = int(m_single.group(1))
        else:
            return None

        m_pct = re.search(r'([\d.]+)\s*%', seg)
        if not m_pct:
            return None

        tiers.append((age_from, age_to, float(m_pct.group(1))))

    return tiers if len(tiers) >= 2 else None


# Header keywords that name a tier/band/discount concept but may not match any
# mapping_engine synonym exactly (e.g. a lone "Slab" or "Discount Band" header).
# Used by _classify_columns to keep such a column's text flowing into the parse
# instead of being dropped as UNKNOWN.
SLAB_HINT_KEYWORDS = [
    "slab", "tier", "band", "discount", "premium slab", "si slab",
]


# Known locations in Indian insurance grids
STATES_LIST = [
    "andaman", "nicobar", "andhra pradesh", "andra pradesh", "ap/ts", "ap", "ts", "tg", "telangana", 
    "telengana", "arunachal", "ar", "assam", "as", "bihar", "bh", "chandigarh", "chhattisgarh", "cg", "dadra", 
    "daman", "diu", "delhi", "ncr", "dl", "goa", "ga", "gujarat", "gj", "haryana", "hr", "himachal", 
    "jammu", "kashmir", "jharkhand", "jh", "karnataka", "ka", "kerala", "kl", "ladakh", 
    "lakshadweep", "madhya pradesh", "mp", "maharashtra", "mh", "manipur", "meghalaya", "ml", 
    "mizoram", "nagaland", "nl", "odisha", "orissa", "puducherry", "pondicherry", "punjab", "pb", 
    "rajasthan", "rj", "sikkim", "tamil nadu", "tn", "tripura", "tr", "uttar pradesh", "up", 
    "uttarakhand", "west bengal", "wb", "pan india", "ahmedabad", "bangalore", "bhubaneshwar", 
    "mumbai", "nagpur", "pune", "surat", "vadodara", "vijaywada", "vishakapatnam", "central odisha",
    "metro"
]

STATE_ABBR_MAP = {
    "ap/ts": "AP, TS",
    "andra pradesh": "AP",
    "andhra pradesh": "AP",
    "telangana": "TS",
    "telengana": "TS",
    "tg": "TS",
    "tamil nadu": "TN",
    "tn": "TN",
    "karnataka": "KA",
    "ka": "KA",
    "maharashtra": "MH",
    "mh": "MH",
    "goa": "GA",
    "ga": "GA",
    "gujarat": "GJ",
    "gj": "GJ",
    "bihar": "BR",
    "bh/jh": "BR, JH",
    "bh": "BR",
    "jh": "JH",
    "jharkhand": "JH",
    "west bengal": "WB",
    "wb": "WB",
    "orissa": "OD",
    "odisha": "OD",
    "chhattisgarh": "CG",
    "cg": "CG",
    "madhya pradesh": "MP",
    "mp": "MP",
    "rajasthan": "RJ",
    "rj": "RJ",
    "uttar pradesh": "UP",
    "up": "UP",
    "delhi": "DL",
    "dl": "DL",
    "punjab": "PB",
    "pb": "PB",
    "haryana": "HR",
    "hr": "HR",
    "kerala": "KL",
    "kl": "KL",
    "assam": "AS",
    "as": "AS",
    "tripura": "TR",
    "tr": "TR",
    "assam/tripura": "AS, TR",
    "meghalaya": "ML",
    "ml": "ML",
    "arunachal": "AR",
    "ar": "AR",
    "nagaland": "NL",
    "nl": "NL",
    "pan india": "ALL"
}

class ExcelParserService:
    def __init__(self):
        self.mapper = ColumnMapper()
        self.normalizer = ValueNormalizer()

    def _get_scalar_value(self, row: pd.Series, col: Any) -> Any:
        if col is None or col not in row:
            return None
        val = row[col]
        if isinstance(val, pd.Series):
            val = val.iloc[0]
        # Automatically serialize dates to YYYY-MM-DD strings to prevent database JSON serialization errors
        if val is not None and hasattr(val, "isoformat"):
            val_str = val.isoformat()
            val = val_str.split("T")[0] if "T" in val_str else val_str
        return val


    def _get_row_header_score(self, row_values: List[Any]) -> int:
        score = 0
        matched_stds = set()
        for val in row_values:
            if val is None or pd.isna(val):
                continue
            val_clean = str(val).strip().lower()
            if val_clean == "":
                continue
                
            matched = False
            for std, syns in self.mapper.mappings.items():
                if val_clean in syns:
                    if std not in matched_stds:
                        score += 5  # High score for parameters
                        matched_stds.add(std)
                    matched = True
                    break
            if not matched:
                # Split header string into words to match locations exactly
                words = [w.strip() for w in val_clean.replace("-", " ").replace("/", " ").replace("(", " ").replace(")", " ").split()]
                for w in words:
                    if w in STATES_LIST:
                        score += 1
                        break
        return score

    def _combine_headers_and_detect_layout(self, df: pd.DataFrame) -> Tuple[List[str], int]:
        # Find the row index with the highest header score
        max_score = -1
        header_idx = 0
        for idx in range(min(10, len(df))):
            row_vals = list(df.iloc[idx].values)
            score = self._get_row_header_score(row_vals)
            if score > max_score:
                max_score = score
                header_idx = idx
                
        df_sliced = df.iloc[header_idx:].reset_index(drop=True)
        num_cols = len(df_sliced.columns)
        
        # Determine the number of header rows
        header_rows_count = 1
        for offset in range(1, 4):
            r_idx = header_idx + offset
            if r_idx >= len(df):
                break
            row_vals = list(df.iloc[r_idx].values)
            # Stop if the row contains numbers (data values)
            num_count = sum(1 for x in row_vals if isinstance(x, (int, float)) and x is not None and not pd.isna(x))
            if num_count > 0:
                header_rows_count = offset
                break
            else:
                header_rows_count = offset + 1

        combined_headers = []
        for col_idx in range(num_cols):
            tokens = []
            for row_idx in range(header_rows_count):
                val = df_sliced.iloc[row_idx, col_idx]
                if val is not None:
                    val_str = str(val).strip()
                    if val_str.lower() not in ("none", "nan", "") and val_str not in tokens:
                        tokens.append(val_str)
            combined = " - ".join(tokens) if tokens else f"Unnamed_{col_idx}"
            combined_headers.append(combined)
            
        return combined_headers, header_idx + header_rows_count

    def _classify_columns(self, headers: List[str]) -> List[Dict[str, Any]]:
        classifications = []
        for idx, h in enumerate(headers):
            h_clean = h.strip().lower()
            words = [w.strip() for w in h_clean.replace("-", " ").replace("/", " ").replace("(", " ").replace(")", " ").split()]

            # Match standard field synonyms using Phase-based approach
            matched_param = None

            # Phase 1: Exact match scan
            for std, syns in self.mapper.mappings.items():
                if h_clean in syns:
                    matched_param = std
                    break

            # Phase 2: Specific substring match (longest matching synonym wins)
            if not matched_param:
                best_syn_len = -1
                for std, syns in self.mapper.mappings.items():
                    for syn in syns:
                        if syn in h_clean and len(syn) > best_syn_len:
                            # Avoid false short word matches (like "type" or "od" matching subwords)
                            if len(syn) <= 4:
                                if syn not in words:
                                    continue
                            best_syn_len = len(syn)
                            matched_param = std

            # Check if this column is a Location Column using exact word matching
            matched_locations = []
            for w in words:
                if w in STATES_LIST:
                    abbr = STATE_ABBR_MAP.get(w, w.upper())
                    for a in abbr.split(","):
                        a_clean = a.strip().upper()
                        if a_clean not in matched_locations:
                            matched_locations.append(a_clean)

            if matched_locations:
                matched_location = ", ".join(matched_locations)
                
                matched_zone = None
                header_lower = h.lower()
                for dir_word in ["east", "west", "north", "south", "central", "rest of", "ro"]:
                    if dir_word in header_lower:
                        clean_dir = dir_word.replace(" ", "").capitalize()
                        prefix = matched_locations[0]
                        matched_zone = f"{prefix}-{clean_dir}"
                        break

                classifications.append({
                    "index": idx,
                    "header": h,
                    "type": "LOCATION",
                    "location": matched_location,
                    "zone": matched_zone
                })
            elif matched_param:
                classifications.append({
                    "index": idx,
                    "header": h,
                    "type": "RATE" if matched_param.startswith("payin_") or matched_param.startswith("payout_") or matched_param.startswith("slab_") or matched_param in ("payin_type", "premium_type") else "PARAM",
                    "field": matched_param
                })
            elif any(kw in h_clean for kw in SLAB_HINT_KEYWORDS):
                # A column that names a tier/band/discount concept but doesn't match
                # any known synonym (e.g. a lone "Slab"/"Discount Band" header) would
                # otherwise be dropped entirely as UNKNOWN — its cell text never even
                # reaching raw_json — leaving slab_from/slab_to with nothing to parse
                # a boundary out of. Route it through the PARAM pipeline under a
                # synthetic field so classify_rule_type/extract_numeric_range can
                # still use its text as a slab-boundary fallback.
                classifications.append({
                    "index": idx,
                    "header": h,
                    "type": "PARAM",
                    "field": "_slab_hint_text"
                })
            else:
                classifications.append({
                    "index": idx,
                    "header": h,
                    "type": "UNKNOWN"
                })
        return classifications

    def _infer_fields_from_text(self, rule_data: Dict[str, Any], raw_json: Dict[str, Any]):
        # Combine all non-empty cell values from the row to construct the context
        text_parts = []
        for k, v in raw_json.items():
            if v is not None and str(v).strip() != "":
                text_parts.append(str(v).strip())
        
        # Also append sheet name, product, subclass, and remarks to have full context
        text_parts.append(rule_data.get("sheet_name") or "")
        text_parts.append(rule_data.get("product") or "")
        text_parts.append(rule_data.get("sub_class") or "")
        text_parts.append(rule_data.get("remarks") or "")
        
        row_text = " | ".join(text_parts)
        row_text_lower = row_text.lower()

        # 1. NCB Status
        if any(kw in row_text_lower for kw in ["without ncb", "no ncb", "ncb exclusion", "ncb excl"]):
            rule_data["ncb_status"] = "No"
        elif any(kw in row_text_lower for kw in ["ncb cases", "ncb only", "ncb >", "ncb >=", "ncb percent", "ncb ="]):
            rule_data["ncb_status"] = "Yes"
        elif "ncb" in row_text_lower:
            rule_data["ncb_status"] = "Yes"

        # 2. CPA Status
        if any(kw in row_text_lower for kw in ["without cpa", "no cpa", "excluding cpa", "cpa excluded"]):
            rule_data["cpa_status"] = "No"
        elif any(kw in row_text_lower for kw in ["cpa cases", "cpa only", "cpa cover", "cpa included", "cpa yes", "cpa y/n", "cpa y"]):
            rule_data["cpa_status"] = "Yes"

        # 3. Fuel Type
        if "petrol" in row_text_lower and not any(x in row_text_lower for x in ["except petrol", "other than petrol"]):
            rule_data["fuel_type"] = "PETROL"
        elif "diesel" in row_text_lower and not any(x in row_text_lower for x in ["except diesel", "other than diesel"]):
            rule_data["fuel_type"] = "DIESEL"
        elif "cng" in row_text_lower:
            rule_data["fuel_type"] = "CNG"
        elif any(kw in row_text_lower for kw in ["electric", "ev", "e-rickshaw", "e-cart"]):
            rule_data["fuel_type"] = "ELECTRIC"

        # 4. Make
        if "except new mahindra" in row_text_lower:
            rule_data["make"] = "EXCEPT NEW MAHINDRA"
        elif "mahindra only" in row_text_lower:
            rule_data["make"] = "MAHINDRA"
        elif "except tata" in row_text_lower:
            rule_data["make"] = "EXCEPT TATA"
        elif "hyundai & maruti" in row_text_lower or "hyundai and maruti" in row_text_lower:
            rule_data["make"] = "HYUNDAI, MARUTI"
        elif "hyundai" in row_text_lower:
            rule_data["make"] = "HYUNDAI"
        elif "maruti" in row_text_lower:
            rule_data["make"] = "MARUTI"
        elif "honda" in row_text_lower:
            rule_data["make"] = "HONDA"
        elif "bajaj" in row_text_lower:
            rule_data["make"] = "BAJAJ"

        # 5. Model
        for model_name in ["nexon", "swift", "alto 800", "alto", "ace", "super ace", "yodha", "xenon", "magic", "intra", "super carry", "jeeto", "supro"]:
            if model_name in row_text_lower:
                rule_data["model"] = model_name.upper()
                break

        # 6. Policy Type / Plan Type Normalization & Inference
        p_type = rule_data.get("policy_type")
        if p_type:
            p_type_lower = str(p_type).lower()
            if any(kw in p_type_lower for kw in ["saod", "standalone od", "sod", "standalone own damage"]):
                rule_data["policy_type"] = "Standalone Own Damage"
            elif any(kw in p_type_lower for kw in ["satp", "third party", "tp only", "act only", "tp policy", "tp cases", "stp"]):
                rule_data["policy_type"] = "Third Party"
            elif any(kw in p_type_lower for kw in ["package", "comprehensive", "p & l", "pkg"]):
                rule_data["policy_type"] = "Comprehensive"
        else:
            if any(kw in row_text_lower for kw in ["saod", "standalone od", "sod", "standalone own damage"]):
                rule_data["policy_type"] = "Standalone Own Damage"
            elif any(kw in row_text_lower for kw in ["satp", "third party", "tp only", "act only", "tp policy", "tp cases"]):
                rule_data["policy_type"] = "Third Party"
            elif any(kw in row_text_lower for kw in ["package", "comprehensive", "p & l", "pkg"]):
                rule_data["policy_type"] = "Comprehensive"

    def parse_workbook(self, file_bytes: bytes, filename: str = None) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=False, data_only=True)
        sheet_names = wb.sheetnames
        
        # Determine fallback company name
        company = "Unknown"
        if filename:
            fn_lower = filename.lower()
            if "tata" in fn_lower:
                company = "Tata"
            elif "shriram" in fn_lower:
                company = "Shriram"
            elif "chola" in fn_lower:
                company = "Cholamandalam"
            elif "digit" in fn_lower:
                company = "Digit"

        print(f"\n[STARTING WORKBOOK PARSING] Sheets detected: {sheet_names}")
        all_parsed_rules = []
        existing_keys = set()  # To detect duplicates across the workbook

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            
            # Skip lookup master sheets for Tata
            sheet_name_lower = sheet_name.lower()
            if "vehicle type master" in sheet_name_lower or ("master" in sheet_name_lower and "vehicle" in sheet_name_lower) or "incremental" in sheet_name_lower:
                print(f"  [SHEET SKIP] Skipping Tata lookup/incremental sheet: '{sheet_name}'.")
                continue

            # Skip reference/lookup-only sheets (RTO/zone code tables, decline
            # lists, terms & conditions) — these carry no commission data at
            # all, they're just cross-reference tables for other sheets.
            LOOKUP_SHEET_KEYWORDS = [
                "zone", "metro rto", "rto codes", "conditions", "declined",
                "decline", "pan india",
            ]
            if any(kw in sheet_name_lower for kw in LOOKUP_SHEET_KEYWORDS):
                print(f"  [SHEET SKIP] '{sheet_name}' is a reference/lookup table, not commission data.")
                continue

            # This app is Motor-only — skip sheets that are clearly a
            # different line of business (e.g. Shriram's "NM grid" covering
            # Fire & Engineering/GPA/Marine/Liability/Co-operative, Go
            # Digit's "Non Motor"/"School and Staff Bus").
            NON_MOTOR_SHEET_KEYWORDS = ["non motor", "nm grid", "school", "staff bus"]
            if any(kw in sheet_name_lower for kw in NON_MOTOR_SHEET_KEYWORDS):
                print(f"  [SHEET SKIP] '{sheet_name}' is non-Motor LOB — this app only extracts Motor data.")
                continue

            # Check if sheet is empty
            if ws.max_row <= 1 and ws.max_column <= 1:
                print(f"  [SHEET SKIP] '{sheet_name}' is empty.")
                continue

            print(f"\n  [PROCESSING SHEET] '{sheet_name}' | Total Rows: {ws.max_row} | Total Columns: {ws.max_column}")

            # Unmerge cells in this sheet and copy values
            merged_ranges = list(ws.merged_cells.ranges)
            if merged_ranges:
                print(f"    Unmerging {len(merged_ranges)} cell ranges...")
            for r in merged_ranges:
                min_col, min_row, max_col, max_row = r.bounds
                top_left_value = ws.cell(row=min_row, column=min_col).value
                try:
                    ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                except ValueError:
                    pass
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        ws.cell(row=row, column=col, value=top_left_value)

            # Load into pandas
            data = list(ws.values)
            if not data:
                print(f"    [SHEET SKIP] No values parsed in '{sheet_name}' after unmerging.")
                continue
                
            df = pd.DataFrame(data)
            
            # Detect combined headers and boundary index
            headers, data_start_idx = self._combine_headers_and_detect_layout(df)
            print(f"    Detected header rows count. Data starts at row index {data_start_idx}.")
            
            # Classify columns
            classifications = self._classify_columns(headers)
            locations = [c for c in classifications if c["type"] == "LOCATION"]
            params = [c for c in classifications if c["type"] == "PARAM"]
            rates = [c for c in classifications if c["type"] == "RATE"]
            
            # Determine Layout: PIVOT if it has multiple location columns
            is_pivot = len(locations) > 3
            layout_type = "PIVOT" if is_pivot else "FLAT"
            print(f"    Sheet layout identified: {layout_type}")
            print(f"    Classified columns: {len(locations)} Location columns, {len(params)} Parameter columns, {len(rates)} Rate columns.")
            
            if not params and not rates and not locations:
                print(f"    [SHEET SKIP] No parameter, rate or location columns could be identified.")
                continue

            # Assign headers as column names to make it readable in row Series
            df_sliced = df.iloc[data_start_idx:].copy()
            df_sliced.columns = headers

            sheet_rules_count = self._process_flat_rows(
                df_sliced, headers, classifications, is_pivot, sheet_name, filename, company,
                existing_keys, all_parsed_rules
            )
            print(f"    [SHEET COMPLETED] Extracted {sheet_rules_count} rules from '{sheet_name}'.")

        final_rules = self._group_and_merge_rules(all_parsed_rules)
        return final_rules

    def parse_table(
        self, headers: List[str], rows: List[List[Any]], sheet_name: str,
        filename: Optional[str], company: str, existing_keys: set,
        all_parsed_rules: List[Dict[str, Any]]
    ) -> int:
        """
        Classifies and parses an already-tabular data source (one row = one rule,
        e.g. a Word table or a PDF table) through the same column-classification and
        normalization pipeline used for Excel FLAT/PIVOT sheets. Appends parsed rule
        dicts into all_parsed_rules (shared across every table/page in the document,
        so the caller can run one document-wide _group_and_merge_rules pass at the end).
        Returns the number of rules extracted from this table.
        """
        if not headers or not rows:
            return 0

        clean_headers = [str(h).strip() if h is not None and str(h).strip() else f"Unnamed_{i}" for i, h in enumerate(headers)]
        df = pd.DataFrame(rows, columns=clean_headers)

        classifications = self._classify_columns(clean_headers)
        locations = [c for c in classifications if c["type"] == "LOCATION"]
        params = [c for c in classifications if c["type"] == "PARAM"]
        rates = [c for c in classifications if c["type"] == "RATE"]

        if not params and not rates and not locations:
            return 0

        is_pivot = len(locations) > 3
        return self._process_flat_rows(
            df, clean_headers, classifications, is_pivot, sheet_name, filename, company,
            existing_keys, all_parsed_rules
        )

    def _process_flat_rows(
        self, df_sliced: pd.DataFrame, headers: List[str], classifications: List[Dict[str, Any]],
        is_pivot: bool, sheet_name: str, filename: Optional[str], company: str,
        existing_keys: set, all_parsed_rules: List[Dict[str, Any]]
    ) -> int:
        """
        Parses each data row of an already-tabular, header-classified dataset into rule dict(s),
        appending them to all_parsed_rules. Shared by Excel FLAT/PIVOT sheets and by the
        Word/PDF table parsers (which are inherently one-row-per-rule, i.e. FLAT).
        """
        locations = [c for c in classifications if c["type"] == "LOCATION"]
        params = [c for c in classifications if c["type"] == "PARAM"]
        rates = [c for c in classifications if c["type"] == "RATE"]
        sheet_rules_count = 0
        for r_idx, row in df_sliced.iterrows():
                # Check ignore
                if self._is_ignored_row(row, headers):
                    continue
                    
                # Build common parameter values
                param_vals = {}
                raw_json = {}
                
                # Initialize fields to None
                for field in ("lob", "file_type", "insurance_company", "product", "policy_type", 
                              "plan_type", "sub_product", "class_", "sub_class", "make", "model", 
                              "fuel_type", "body_type", "cpa_status", "ncb_status", "partner_type", 
                              "zone", "source", "rto", "remarks", "state", "vehicle_age", "effective_date",
                              "vehicle_age_from", "vehicle_age_to"):
                    param_vals[field] = None

                for p in params:
                    val = self._get_scalar_value(row, p["header"])
                    raw_json[p["header"]] = val
                    if val is not None and str(val).strip() != "":
                        # Only set value if not already set, or overwrite if None
                        if param_vals.get(p["field"]) is None:
                            param_vals[p["field"]] = str(val).strip()
                
                # Simple standard fields mapping
                base_rule = {}
                for field in ("lob", "file_type", "insurance_company", "product", "policy_type", 
                              "plan_type", "sub_product", "class_", "sub_class", "make", "model", 
                              "fuel_type", "body_type", "cpa_status", "ncb_status", "partner_type", 
                              "zone", "source", "rto", "remarks"):
                    source_field = "class_" if field == "class_" else field
                    base_rule[field] = param_vals.get(source_field)
                
                # Normalization Rule: product range splitting
                product_str = base_rule.get("product")
                if product_str:
                    norm_prod, extracted_subclass = split_product_range(product_str)
                    if extracted_subclass:
                        base_rule["product"] = norm_prod
                        current_subclass = base_rule.get("sub_class")
                        if current_subclass and str(current_subclass).strip() != "":
                            base_rule["sub_class"] = f"{current_subclass} {extracted_subclass}".strip()
                        else:
                            base_rule["sub_class"] = extracted_subclass

                # Populate default fallback company
                if not base_rule["insurance_company"]:
                    base_rule["insurance_company"] = company

                # Special normalizations
                # 1. State (for FLAT layouts)
                state_val = param_vals.get("state")
                base_rule["state"] = self.normalizer.normalize_states(state_val) if state_val is not None else None
                
                # 2. Vehicle age
                age_val = param_vals.get("vehicle_age")
                if age_val is not None:
                    age_from, age_to = self.normalizer.normalize_vehicle_age(age_val)
                    base_rule["vehicle_age_from"] = age_from
                    base_rule["vehicle_age_to"] = age_to
                else:
                    age_from_val = param_vals.get("vehicle_age_from")
                    age_to_val = param_vals.get("vehicle_age_to")
                    if age_from_val is not None or age_to_val is not None:
                        try:
                            base_rule["vehicle_age_from"] = int(float(age_from_val)) if age_from_val is not None else 1
                        except ValueError:
                            base_rule["vehicle_age_from"] = 1
                        try:
                            base_rule["vehicle_age_to"] = int(float(age_to_val)) if age_to_val is not None else 50
                        except ValueError:
                            base_rule["vehicle_age_to"] = 50
                    else:
                        base_rule["vehicle_age_from"] = None
                        base_rule["vehicle_age_to"] = None
                    
                # 3. Effective date (Auto-corrected to today's date if missing)
                date_val = param_vals.get("effective_date")
                parsed_dt = self.normalizer.normalize_date(date_val) if date_val is not None else None
                base_rule["effective_date"] = parsed_dt if parsed_dt else date.today()
                
                # Fallback LOB detection
                if not base_rule["lob"]:
                    check_str = f"{sheet_name} {filename or ''} {base_rule['product'] or ''}".lower()
                    if any(x in check_str for x in ["gcv", "pccv", "gccv", "pcv", "motor", "wheeler", "tw", "car", "moped", "scooter", "bike", "private"]):
                        base_rule["lob"] = "Motor"
                    elif any(x in check_str for x in ["health", "medical", "mediclaim", "care"]):
                        base_rule["lob"] = "Health"
                    elif any(x in check_str for x in ["life", "term", "endowment"]):
                        base_rule["lob"] = "Life"

                # Read flat rates if any
                flat_rates = {}
                for r in rates:
                    val = self._get_scalar_value(row, r["header"])
                    flat_rates[r["field"]] = val
                    raw_json[r["header"]] = val

                # Capture UNKNOWN column text too — a column with no recognized
                # synonym (a free-text "Remarks"/"Condition"/notes column, e.g.
                # "NCB Cases", "Except New Mahindra", "HR 68 Excluded") previously
                # never reached raw_json at all, so downstream inference (Groq
                # enrichment) had nothing to read for exactly the descriptive text
                # it needs. This is purely additive — doesn't change classification
                # or any existing field mapping.
                for u in [c for c in classifications if c["type"] == "UNKNOWN"]:
                    val = self._get_scalar_value(row, u["header"])
                    if val is not None and str(val).strip() != "":
                        raw_json[u["header"]] = val

                if is_pivot:
                    # For pivot tables, each location column is a rate entry
                    for loc in locations:
                        rate_val = self._get_scalar_value(row, loc["header"])
                        raw_json[loc["header"]] = rate_val
                        
                        # Skip empty, null, or zero rates to prevent clutter
                        if rate_val is None or pd.isna(rate_val) or str(rate_val).strip() in ("", "0", "0%"):
                            continue
                            
                        # Build a separate rule record for this location
                        rule_data = base_rule.copy()
                        rule_data["sheet_name"] = sheet_name
                        rule_data["state"] = loc["location"]
                        if loc.get("zone"):
                            rule_data["zone"] = loc["zone"]
                        
                        # Heuristics to determine whether this location column represents OD or TP
                        header_lower = loc["header"].lower()
                        rate_field = "payin_net"  # Default
                        if "comp" in header_lower or "pack" in header_lower or "saod" in header_lower or "od" in header_lower:
                            rate_field = "payin_od"
                        elif "tp" in header_lower or "satp" in header_lower or "act" in header_lower:
                            rate_field = "payin_tp"
                            
                        # If a policy type column was mapped, use it. Otherwise, set default from rate field
                        if not rule_data["policy_type"]:
                            rule_data["policy_type"] = "Comprehensive" if rate_field == "payin_od" else "Third Party"
                            
                        # Run text based field inference to scan remarks, headers, and column strings
                        self._infer_fields_from_text(rule_data, raw_json)

                        # Detect vehicle-age-tiered text packed into this single cell
                        # (e.g. Go Digit's "Age 0: 60%/85%\nAge >=1: 85%"). A cell with
                        # >=2 distinct age bands and differing rates is a genuine slab
                        # (Vehicle Age Slab), keyed by age instead of premium/volume —
                        # emitted as ONE record with a SlabDetail per age tier, rather
                        # than N separate flat rows (which would silently drop the rate
                        # to null on a single-value parse, and — per the current CRM
                        # normalization spec — age-tiered cells are SLAB, not NON_SLAB).
                        age_tiers = parse_tiered_rate_text(rate_val)

                        if age_tiers is not None and len({pct for _, _, pct in age_tiers}) >= 2:
                            tier_rule_data = rule_data.copy()

                            status, warnings = RuleValidator.validate_rule(tier_rule_data, existing_keys)
                            tier_rule_data["raw_json"] = raw_json.copy()
                            tier_rule_data["validation_status"] = status
                            tier_rule_data["warnings"] = warnings

                            tier_rule_data["commission_type"] = "SLAB"
                            tier_rule_data["slab_configuration"] = True
                            tier_rule_data["payin_od"] = None
                            tier_rule_data["payout_od"] = None
                            tier_rule_data["payin_tp"] = None
                            tier_rule_data["payout_tp"] = None
                            tier_rule_data["payin_net"] = None
                            tier_rule_data["payout_net"] = None
                            tier_rule_data["payin_reward"] = None
                            tier_rule_data["payout_reward"] = None
                            tier_rule_data["payin_scheme"] = None
                            tier_rule_data["payout_scheme"] = None

                            tier_rule_data["slabs"] = [
                                {
                                    "payin_type": "PERCENTAGE",
                                    "premium_type": {"payin_od": "OD", "payin_tp": "TP"}.get(rate_field, "NET"),
                                    "slab_from": age_from,
                                    "slab_to": age_to,
                                    "payin_od": pct if rate_field == "payin_od" else None,
                                    "payout_od": None,
                                    "payin_tp": pct if rate_field == "payin_tp" else None,
                                    "payout_tp": None,
                                    "payin_net": pct if rate_field == "payin_net" else None,
                                    "payout_net": None,
                                }
                                for age_from, age_to, pct in age_tiers
                            ]
                            all_parsed_rules.append(tier_rule_data)
                            sheet_rules_count += 1
                        else:
                            rate_percent = self.normalizer.normalize_percentage(rate_val)
                            tier_rule_data = rule_data.copy()

                            status, warnings = RuleValidator.validate_rule(tier_rule_data, existing_keys)
                            tier_rule_data["raw_json"] = raw_json.copy()
                            tier_rule_data["validation_status"] = status
                            tier_rule_data["warnings"] = warnings

                            tier_rule_data["commission_type"] = "NON_SLAB"
                            tier_rule_data["slab_configuration"] = False

                            tier_rule_data["payin_od"] = rate_percent if rate_field == "payin_od" else None
                            tier_rule_data["payout_od"] = None
                            tier_rule_data["payin_tp"] = rate_percent if rate_field == "payin_tp" else None
                            tier_rule_data["payout_tp"] = None
                            tier_rule_data["payin_net"] = rate_percent if rate_field == "payin_net" else None
                            tier_rule_data["payout_net"] = None
                            tier_rule_data["payin_reward"] = None
                            tier_rule_data["payout_reward"] = None
                            tier_rule_data["payin_scheme"] = None
                            tier_rule_data["payout_scheme"] = None

                            tier_rule_data["slabs"] = []
                            all_parsed_rules.append(tier_rule_data)
                            sheet_rules_count += 1
                else:
                    # Flat layouts: single rule record
                    rule_data = base_rule.copy()
                    rule_data["sheet_name"] = sheet_name
                    
                    # Run text based field inference to scan remarks, headers, and column strings
                    self._infer_fields_from_text(rule_data, raw_json)

                    # Slabs extraction from row rates
                    payin_od = self.normalizer.normalize_percentage(flat_rates.get("payin_od")) if flat_rates.get("payin_od") is not None else None
                    payout_od = self.normalizer.normalize_percentage(flat_rates.get("payout_od")) if flat_rates.get("payout_od") is not None else None
                    payin_tp = self.normalizer.normalize_percentage(flat_rates.get("payin_tp")) if flat_rates.get("payin_tp") is not None else None
                    payout_tp = self.normalizer.normalize_percentage(flat_rates.get("payout_tp")) if flat_rates.get("payout_tp") is not None else None
                    payin_net = self.normalizer.normalize_percentage(flat_rates.get("payin_net")) if flat_rates.get("payin_net") is not None else None
                    payout_net = self.normalizer.normalize_percentage(flat_rates.get("payout_net")) if flat_rates.get("payout_net") is not None else None
                    payin_reward = self.normalizer.normalize_percentage(flat_rates.get("payin_reward")) if flat_rates.get("payin_reward") is not None else None
                    payout_reward = self.normalizer.normalize_percentage(flat_rates.get("payout_reward")) if flat_rates.get("payout_reward") is not None else None
                    payin_scheme = self.normalizer.normalize_percentage(flat_rates.get("payin_scheme")) if flat_rates.get("payin_scheme") is not None else None
                    payout_scheme = self.normalizer.normalize_percentage(flat_rates.get("payout_scheme")) if flat_rates.get("payout_scheme") is not None else None
                    
                    # Slab detection
                    is_slab = classify_rule_type(sheet_name, headers, rule_data, flat_rates) == "SLAB"
                    
                    # Validate
                    status, warnings = RuleValidator.validate_rule(rule_data, existing_keys)
                    
                    rule_data["raw_json"] = raw_json
                    rule_data["validation_status"] = status
                    rule_data["warnings"] = warnings
                    
                    if is_slab:
                        rule_data["commission_type"] = "SLAB"
                        rule_data["slab_configuration"] = True
                        
                        slab_from = self.normalizer.normalize_percentage(flat_rates.get("slab_from")) if flat_rates.get("slab_from") is not None else None
                        slab_to = self.normalizer.normalize_percentage(flat_rates.get("slab_to")) if flat_rates.get("slab_to") is not None else None
                        
                        if slab_from is None and slab_to is None:
                            slab_from, slab_to = extract_numeric_range(rule_data.get("sub_class") or rule_data.get("product"))
                        if slab_from is None and slab_to is None:
                            # Last resort: a column that named a tier/band/discount
                            # concept but wasn't a recognized synonym (see
                            # SLAB_HINT_KEYWORDS / _classify_columns).
                            slab_from, slab_to = extract_numeric_range(param_vals.get("_slab_hint_text"))

                        if slab_from is not None and (slab_from == 0.0 or slab_from == 0):
                            slab_from = 1.0

                        payin_type = str(flat_rates.get("payin_type")).strip() if flat_rates.get("payin_type") is not None else "PERCENTAGE"
                        premium_type = str(flat_rates.get("premium_type")).strip() if flat_rates.get("premium_type") is not None else None
                        
                        slab_detail = {
                            "payin_type": payin_type,
                            "premium_type": premium_type,
                            "slab_from": slab_from,
                            "slab_to": slab_to,
                            "payin_od": payin_od,
                            "payout_od": payout_od,
                            "payin_tp": payin_tp,
                            "payout_tp": payout_tp,
                            "payin_net": payin_net,
                            "payout_net": payout_net
                        }
                        rule_data["slabs"] = [slab_detail]
                        
                        rule_data["payin_od"] = None
                        rule_data["payout_od"] = None
                        rule_data["payin_tp"] = None
                        rule_data["payout_tp"] = None
                        rule_data["payin_net"] = None
                        rule_data["payout_net"] = None
                        rule_data["payin_reward"] = None
                        rule_data["payout_reward"] = None
                        rule_data["payin_scheme"] = None
                        rule_data["payout_scheme"] = None
                    else:
                        rule_data["commission_type"] = "NON_SLAB"
                        rule_data["slab_configuration"] = False
                        rule_data["slabs"] = []
                        
                        rule_data["payin_od"] = payin_od
                        rule_data["payout_od"] = payout_od
                        rule_data["payin_tp"] = payin_tp
                        rule_data["payout_tp"] = payout_tp
                        rule_data["payin_net"] = payin_net
                        rule_data["payout_net"] = payout_net
                        rule_data["payin_reward"] = payin_reward
                        rule_data["payout_reward"] = payout_reward
                        rule_data["payin_scheme"] = payin_scheme
                        rule_data["payout_scheme"] = payout_scheme
                        
                    all_parsed_rules.append(rule_data)
                    sheet_rules_count += 1

        return sheet_rules_count

    def _group_and_merge_rules(self, all_parsed_rules: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Collapses business-identical rows into one, and folds genuinely-differing
        rate rows sharing the same business key into a single SLAB rule with
        nested tiers. Shared by Excel workbook parsing and the Word/PDF parsers.
        """
        # --- GROUP & MERGE RULES ENGINE ---
        grouped_rules: Dict[Tuple, List[Dict[str, Any]]] = {}
        for r in all_parsed_rules:
            key = (
                str(r.get("lob") or "").strip().upper(),
                str(r.get("file_type") or "").strip().upper(),
                str(r.get("insurance_company") or "").strip().upper(),
                str(r.get("product") or "").strip().upper(),
                str(r.get("policy_type") or "").strip().upper(),
                str(r.get("plan_type") or "").strip().upper(),
                str(r.get("sub_product") or "").strip().upper(),
                str(r.get("class") or "").strip().upper(),
                str(r.get("sub_class") or "").strip().upper(),
                str(r.get("make") or "").strip().upper(),
                str(r.get("model") or "").strip().upper(),
                str(r.get("fuel_type") or "").strip().upper(),
                str(r.get("body_type") or "").strip().upper(),
                str(r.get("vehicle_age_from") or "").strip(),
                str(r.get("vehicle_age_to") or "").strip(),
                str(r.get("cpa_status") or "").strip().upper(),
                str(r.get("ncb_status") or "").strip().upper(),
                str(r.get("partner_type") or "").strip().upper(),
                str(r.get("state") or "").strip().upper(),
                str(r.get("zone") or "").strip().upper(),
                str(r.get("source") or "").strip().upper(),
                str(r.get("rto") or "").strip().upper(),
                str(r.get("remarks") or "").strip().upper(),
                str(r.get("effective_date") or "").strip(),
            )
            grouped_rules.setdefault(key, []).append(r)

        def _rate_signature(r: Dict[str, Any]) -> Tuple:
            if r.get("slabs"):
                return tuple(
                    (s.get("payin_od"), s.get("payin_tp"), s.get("payin_net")) for s in r["slabs"]
                )
            return (
                r.get("payin_od"), r.get("payout_od"), r.get("payin_tp"), r.get("payout_tp"),
                r.get("payin_net"), r.get("payout_net"), r.get("payin_reward"), r.get("payout_reward"),
                r.get("payin_scheme"), r.get("payout_scheme"),
            )

        final_rules = []
        for key, group in grouped_rules.items():
            if len(group) == 1:
                final_rules.append(group[0])
                continue

            # Only a genuine rate difference across otherwise-identical rows
            # indicates a real slab table. Identical rates mean these rows
            # are duplicates/parsing artifacts (e.g. the same row appearing
            # twice) — collapse them to one row instead of manufacturing a
            # fake multi-tier slab out of duplicate data.
            if len({_rate_signature(r) for r in group}) == 1:
                final_rules.append(group[0])
            else:
                base_rule = group[0].copy()
                base_rule["commission_type"] = "SLAB"
                base_rule["slab_configuration"] = True
                
                merged_slabs = []
                for rule_item in group:
                    if rule_item.get("slabs"):
                        for existing_slab in rule_item["slabs"]:
                            if existing_slab.get("slab_from") is None and existing_slab.get("slab_to") is None:
                                # This tier's boundary never got captured upstream —
                                # retry the same text fallbacks used for the
                                # no-slabs-yet branch below, instead of silently
                                # carrying forward a null/null tier (which the
                                # serializer would otherwise paint with an
                                # identical placeholder range for every such tier).
                                retry_from, retry_to = extract_numeric_range(
                                    rule_item.get("sub_class") or rule_item.get("product")
                                )
                                if retry_from is None and retry_to is None:
                                    # Only consider raw_json entries whose ORIGINAL
                                    # COLUMN HEADER looks slab-related — retrying
                                    # against arbitrary columns (a date, an Sr No)
                                    # would misparse them into a bogus range, which
                                    # is worse than leaving the tier boundary null.
                                    raw = rule_item.get("raw_json") or {}
                                    for header_text, v in raw.items():
                                        if not any(kw in str(header_text).lower() for kw in SLAB_HINT_KEYWORDS):
                                            continue
                                        retry_from, retry_to = extract_numeric_range(v)
                                        if retry_from is not None or retry_to is not None:
                                            break
                                existing_slab["slab_from"] = retry_from
                                existing_slab["slab_to"] = retry_to
                        merged_slabs.extend(rule_item["slabs"])
                    else:
                        slab_obj = {
                            "payin_type": "PERCENTAGE",
                            "premium_type": None,
                            "slab_from": None,
                            "slab_to": None,
                            "payin_od": rule_item.get("payin_od"),
                            "payout_od": rule_item.get("payout_od"),
                            "payin_tp": rule_item.get("payin_tp"),
                            "payout_tp": rule_item.get("payout_tp"),
                            "payin_net": rule_item.get("payin_net"),
                            "payout_net": rule_item.get("payout_net")
                        }
                        # Backfill ranges if possible
                        slab_from, slab_to = extract_numeric_range(rule_item.get("sub_class") or rule_item.get("product"))
                        slab_obj["slab_from"] = slab_from
                        slab_obj["slab_to"] = slab_to
                        merged_slabs.append(slab_obj)
                
                # Normalize and deduplicate slabs
                for s in merged_slabs:
                    if s.get("slab_from") is not None and (s["slab_from"] == 0.0 or s["slab_from"] == 0):
                        s["slab_from"] = 1.0

                unique_slabs = []
                seen_slab_keys = set()
                for s in merged_slabs:
                    s_key = (
                        s.get("slab_from"),
                        s.get("slab_to"),
                        s.get("payin_od"),
                        s.get("payin_tp"),
                        s.get("payin_net"),
                    )
                    if s_key not in seen_slab_keys:
                        seen_slab_keys.add(s_key)
                        unique_slabs.append(s)
                merged_slabs = unique_slabs
                
                base_rule["slabs"] = merged_slabs
                dup_warnings = check_duplicate_slab_ranges(merged_slabs)
                if dup_warnings:
                    base_rule["warnings"] = list(base_rule.get("warnings") or []) + dup_warnings
                    base_rule["validation_status"] = "WARNING"

                base_rule["payin_od"] = None
                base_rule["payout_od"] = None
                base_rule["payin_tp"] = None
                base_rule["payout_tp"] = None
                base_rule["payin_net"] = None
                base_rule["payout_net"] = None
                base_rule["payin_reward"] = None
                base_rule["payout_reward"] = None
                base_rule["payin_scheme"] = None
                base_rule["payout_scheme"] = None
                
                final_rules.append(base_rule)
                
        # Post-process: ensure all slab_from values equal to 0.0 or 0 are normalized to 1.0
        for r in final_rules:
            if r.get("slabs"):
                for s in r["slabs"]:
                    if s.get("slab_from") is not None and (s["slab_from"] == 0.0 or s["slab_from"] == 0):
                        s["slab_from"] = 1.0
                        
        print(f"[WORKBOOK PARSING COMPLETED] Grouped {len(all_parsed_rules)} rules into {len(final_rules)} merged rules.")
        return final_rules

    def _is_ignored_row(self, row: pd.Series, headers: List[str]) -> bool:
        if row.isnull().all():
            return True
            
        row_text = " ".join([str(val).lower() for val in row.values if val is not None])
        ignored_keywords = ["total", "grand total", "summary", "notes:", "note:", "*", "disclaimer"]
        
        # Ignore rows with high missing cells and summary keywords
        empty_ratio = row.isnull().sum() / len(row)
        if empty_ratio > 0.6:
            for kw in ignored_keywords:
                if kw in row_text:
                    return True
                    
        # Check if first cell starts with common total markers
        first_val = str(row.iloc[0]).strip().lower() if len(row) > 0 and row.iloc[0] is not None else ""
        if first_val in ("total", "grand total", "summary", "notes"):
            return True
            
        return False
