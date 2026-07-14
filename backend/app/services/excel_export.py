from typing import Any, Dict, List
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from backend.app.models.commission_rule import CommissionRule
from backend.app.services.rule_serializer import serialize_commission_rule

# Column order shared by both sheets — matches the field-by-field spec this
# app normalizes every rule against (see field_normalizer.py/rule_serializer.py).
BUSINESS_COLUMNS: List[tuple] = [
    ("LOB", "lob"),
    ("File Type", "file_type"),
    ("Insurance Company", "insurance_company"),
    ("Product", "product"),
    ("Policy Type", "policy_type"),
    ("Plan Type", "plan_type"),
    ("Sub Product", "sub_product"),
    ("Class", "class"),
    ("Sub Class", "sub_class"),
    ("Make", "make"),
    ("Model", "model"),
    ("Fuel Type", "fuel_type"),
    ("Body Type", "body_type"),
    ("Vehicle Age From", "vehicle_age_from"),
    ("Vehicle Age To", "vehicle_age_to"),
    ("CPA Status", "cpa_status"),
    ("NCB Status", "ncb_status"),
    ("Partner Type", "partner_type"),
    ("State", "state"),
    ("Zone", "zone"),
    ("Source", "source"),
    ("RTO", "rto"),
    ("Effective Date", "effective_date"),
    ("Remarks", "remarks"),
]

NON_SLAB_RATE_COLUMNS: List[tuple] = [
    ("Pay-In OD", "payin_od"), ("Pay-Out OD", "payout_od"),
    ("Pay-In TP", "payin_tp"), ("Pay-Out TP", "payout_tp"),
    ("Pay-In Net", "payin_net"), ("Pay-Out Net", "payout_net"),
    ("Pay-In Reward", "payin_reward"), ("Pay-Out Reward", "payout_reward"),
    ("Pay-In Scheme", "payin_scheme"), ("Pay-Out Scheme", "payout_scheme"),
]

SLAB_TIER_COLUMNS: List[tuple] = [
    ("Pay-In Type", "payin_type"), ("Premium Type", "premium_type"),
    ("Slab From", "slab_from"), ("Slab Upto", "slab_to"),
    ("Pay-In OD", "payin_od"), ("Pay-Out OD", "payout_od"),
    ("Pay-In TP", "payin_tp"), ("Pay-Out TP", "payout_tp"),
    ("Pay-In Net", "payin_net"), ("Pay-Out Net", "payout_net"),
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
DEFAULTED_FONT = Font(color="6D28D9")  # purple-700 — flags a business-default value, not a real extracted one
PARENT_ROW_FONT = Font(bold=True)
PARENT_ROW_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


def _style_header_row(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _autosize_columns(ws) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def build_export_workbook(rules: List[CommissionRule], db: Session) -> BytesIO:
    """
    Generates a two-sheet .xlsx (Non-Slab, Slab) from the FULL filtered set of
    CommissionRule ORM objects — not the paginated page the browser happens to
    have loaded. Reuses serialize_commission_rule so exported values match
    exactly what the UI shows (including business defaults like "ALL"/"NA"),
    and flags any defaulted (non-real) cell in orange via `_defaulted_fields`
    the same serializer already reports.
    """
    non_slab_rules = [r for r in rules if r.commission_type != "SLAB"]
    slab_rules = [r for r in rules if r.commission_type == "SLAB"]

    wb = Workbook()

    # --- NON-SLAB SHEET ---
    ws_ns = wb.active
    ws_ns.title = "Non-Slab"
    ws_ns.append([label for label, _ in BUSINESS_COLUMNS] + [label for label, _ in NON_SLAB_RATE_COLUMNS])
    _style_header_row(ws_ns)
    
    # Freeze header row
    ws_ns.freeze_panes = "A2"
    
    for rule in non_slab_rules:
        data = serialize_commission_rule(rule, db)
        defaulted = set(data.get("_defaulted_fields") or [])
        row_values = [data.get(field) for _, field in BUSINESS_COLUMNS] + [data.get(field) for _, field in NON_SLAB_RATE_COLUMNS]
        ws_ns.append(row_values)
        excel_row = ws_ns.max_row
        for col_idx, (_, field) in enumerate(BUSINESS_COLUMNS, start=1):
            if field in defaulted:
                ws_ns.cell(row=excel_row, column=col_idx).font = DEFAULTED_FONT

    # Apply thin borders to all Non-Slab cells
    for row in ws_ns.iter_rows(min_row=1, max_row=ws_ns.max_row, min_col=1, max_col=ws_ns.max_column):
        for cell in row:
            cell.border = THIN_BORDER

    _autosize_columns(ws_ns)

    # --- SLAB SHEET ---
    ws_sl = wb.create_sheet("Slab")
    
    # Freeze header rows (top 3 rows)
    ws_sl.freeze_panes = "A4"

    # 1. Pre-serialize slab rules to determine max slabs count
    serialized_slab_rules = []
    max_slabs = 0
    for r in slab_rules:
        data = serialize_commission_rule(r, db)
        serialized_slab_rules.append(data)
        slabs = data.get("slabs") or []
        if len(slabs) > max_slabs:
            max_slabs = len(slabs)
            
    # Guarantee at least 1 slab set of headers if there are no rules or max_slabs is 0
    if max_slabs == 0:
        max_slabs = 1

    business_col_count = len(BUSINESS_COLUMNS)
    tier_col_count = len(SLAB_TIER_COLUMNS)
    total_cols = business_col_count + max_slabs * tier_col_count

    # Row 1 (Header level 1)
    row1 = [label for label, _ in BUSINESS_COLUMNS] + [None] * (max_slabs * tier_col_count)
    row1[business_col_count] = "SLAB STRUCTURE"
    ws_sl.append(row1)

    # Row 2 (Header level 2)
    row2 = [None] * business_col_count
    for i in range(1, max_slabs + 1):
        row2.extend([f"Tier {i}"] + [None] * (tier_col_count - 1))
    ws_sl.append(row2)

    # Row 3 (Header level 3)
    row3 = [None] * business_col_count
    for i in range(1, max_slabs + 1):
        row3.extend([label for label, _ in SLAB_TIER_COLUMNS])
    ws_sl.append(row3)

    # Perform Merges
    # Merge Row 1 Slab Structure
    ws_sl.merge_cells(
        start_row=1, start_column=business_col_count + 1,
        end_row=1, end_column=total_cols
    )
    # Merge Row 2 Tiers
    for i in range(1, max_slabs + 1):
        start_col = business_col_count + (i - 1) * tier_col_count + 1
        ws_sl.merge_cells(
            start_row=2, start_column=start_col,
            end_row=2, end_column=start_col + tier_col_count - 1
        )
    # Merge Business Columns Vertically (Rows 1 to 3)
    for col_idx in range(1, business_col_count + 1):
        ws_sl.merge_cells(
            start_row=1, start_column=col_idx,
            end_row=3, end_column=col_idx
        )

    # Style Header cells in rows 1, 2, 3
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx in range(1, 4):
        for col_idx in range(1, total_cols + 1):
            cell = ws_sl.cell(row=r_idx, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = center_align

    # 3. Write rules as single flat rows (starting at row 4)
    for data in serialized_slab_rules:
        defaulted = set(data.get("_defaulted_fields") or [])
        business_values = [data.get(field) for _, field in BUSINESS_COLUMNS]
        
        row_values = business_values.copy()
        slabs = data.get("slabs") or []
        for i in range(max_slabs):
            if i < len(slabs):
                slab = slabs[i]
                slab_values = [slab.get(field) for _, field in SLAB_TIER_COLUMNS]
                row_values.extend(slab_values)
            else:
                row_values.extend([None] * tier_col_count)
                
        ws_sl.append(row_values)
        excel_row = ws_sl.max_row
        
        # Style business column cells if defaulted
        for col_idx, (_, field) in enumerate(BUSINESS_COLUMNS, start=1):
            if field in defaulted:
                ws_sl.cell(row=excel_row, column=col_idx).font = DEFAULTED_FONT
                
        # Style slab columns cells if defaulted
        for i in range(max_slabs):
            if i < len(slabs):
                slab = slabs[i]
                slab_defaulted = set(slab.get("_defaulted_fields") or [])
                start_col = business_col_count + i * tier_col_count + 1
                for col_idx, (_, field) in enumerate(SLAB_TIER_COLUMNS, start=start_col):
                    if field in slab_defaulted:
                        ws_sl.cell(row=excel_row, column=col_idx).font = DEFAULTED_FONT

    # Apply thin borders to all cells
    for row in ws_sl.iter_rows(min_row=1, max_row=ws_sl.max_row, min_col=1, max_col=ws_sl.max_column):
        for cell in row:
            cell.border = THIN_BORDER

    _autosize_columns(ws_sl)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
