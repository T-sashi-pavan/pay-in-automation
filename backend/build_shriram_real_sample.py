import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

OUT_PATH = r"c:\Desktop\ALGONOX\PAY-IN-AUTOMATION\CRM_Ready_Shriram_RealData.xlsx"

NON_SLAB_HEADERS = [
    "LOB", "File Type", "Insurance Company", "Product", "Policy Type", "Plan Type", 
    "Sub Product", "Class", "Sub Class", "Make", "Model", "Fuel Type", "CPA Status", 
    "NCB Status", "Vehicle Age From", "Vehicle Age To", "Partner Type", "Source", 
    "Zone", "State", "RTO", "Effective Date", "Remarks", "Payin OD", "Payin TP", 
    "Payin Net", "Payout OD", "Payout TP", "Payout Net", "Payin Reward", 
    "Payout Reward", "Payin Scheme", "Payout Scheme"
]

SLAB_HEADERS = [
    "LOB", "File Type", "Insurance Company", "Product", "Policy Type", "Plan Type", 
    "Sub Product", "Class", "Sub Class", "Make", "Model", "Fuel Type", "CPA Status", 
    "NCB Status", "Vehicle Age From", "Vehicle Age To", "Partner Type", "Source", 
    "Zone", "State", "RTO", "Effective Date", "Remarks", "Payin Type", "Premium Type", 
    "Payin Slab From", "Payin Slab To", "Payin OD", "Payin TP", "Payin Net", 
    "Payout OD", "Payout TP", "Payout Net"
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

def autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 60)

def split_slabs_by_premium_type(slabs):
    new_slabs = []
    for slab in slabs:
        od = slab.get("Payin OD") or 0.0
        tp = slab.get("Payin TP") or 0.0
        net = slab.get("Payin Net") or 0.0
        
        # Determine which fields are positive
        positive_types = []
        if od > 0.0:
            positive_types.append("OD")
        if tp > 0.0:
            positive_types.append("TP")
        if net > 0.0:
            positive_types.append("NET")
            
        if len(positive_types) > 1:
            for t in positive_types:
                new_slabs.append({
                    "Payin Type": slab.get("Payin Type"),
                    "Premium Type": t,
                    "Payin Slab From": slab.get("Payin Slab From"),
                    "Payin Slab To": slab.get("Payin Slab To"),
                    "Payin OD": od if t == "OD" else 0.0,
                    "Payin TP": tp if t == "TP" else 0.0,
                    "Payin Net": net if t == "NET" else 0.0
                })
        else:
            t = positive_types[0] if positive_types else (slab.get("Premium Type") or "NET")
            new_slabs.append({
                "Payin Type": slab.get("Payin Type"),
                "Premium Type": t,
                "Payin Slab From": slab.get("Payin Slab From"),
                "Payin Slab To": slab.get("Payin Slab To"),
                "Payin OD": od if t == "OD" else 0.0,
                "Payin TP": tp if t == "TP" else 0.0,
                "Payin Net": net if t == "NET" else 0.0
            })
    return new_slabs

STATE_ZONE_RTO_MAP = {
    "AP": {"zone": "Zone B", "rto": "AP-02"},
    "TN": {"zone": "Zone A", "rto": "TN-01"},
    "KA": {"zone": "Zone A", "rto": "KA-01"},
    "MH": {"zone": "Zone A", "rto": "MH-12"},
    "GJ": {"zone": "Zone B", "rto": "GJ-01"},
    "JH": {"zone": "Zone C", "rto": "JH-01"},
    "BR": {"zone": "Zone C", "rto": "BR-01"},
    "WB": {"zone": "Zone B", "rto": "WB-01"},
    "DL": {"zone": "Zone A", "rto": "DL-01"},
    "PB": {"zone": "Zone B", "rto": "PB-02"},
    "AS": {"zone": "Zone C", "rto": "AS-01"},
    "TR": {"zone": "Zone C", "rto": "TR-01"},
    "AR": {"zone": "Zone C", "rto": "AR-01"},
    "ML": {"zone": "Zone C", "rto": "ML-05"},
    "MZ": {"zone": "Zone C", "rto": "MZ-01"},
    "AN": {"zone": "Zone C", "rto": "AN-01"},
    "OR": {"zone": "Zone C", "rto": "OD-02"},
    "ALL": {"zone": "Zone A", "rto": "DL-01"}
}

PARTNER_TYPES_POOL = ["POSP", "MISP", "Broker", "OEM Dealer", "Direct Retail"]
SOURCES_POOL = ["online", "offline"]

def fill_with_realistic_value(h, val, row_dict):
    if h == "Vehicle Age From":
        to_val = row_dict.get("Vehicle Age To")
        if (val == 1 or val == "1") and (to_val == 1 or to_val == "1"):
            return 0

    if val not in [None, "", "ALL", "ALL partners", "NA", "all"]:
        return val
        
    state = row_dict.get("State") or row_dict.get("state") or "ALL"
    state_first = state.split(",")[0].strip()
    geo = STATE_ZONE_RTO_MAP.get(state_first, STATE_ZONE_RTO_MAP["ALL"])
    
    product = str(row_dict.get("Product") or "").lower()
    
    if h == "Make":
        if "2w" in product or "tw" in product or "two wheeler" in product:
            return random.choice(["HERO", "BAJAJ", "RE", "HONDA"])
        elif "gcv" in product or "lcv" in product or "hcv" in product:
            return random.choice(["TATA", "MAHINDRA", "ASHOK LEYLAND"])
        return random.choice(["MARUTI", "HYUNDAI", "TATA", "HONDA"])
        
    elif h == "Model":
        if "2w" in product or "tw" in product or "two wheeler" in product:
            return random.choice(["ACTIVA", "SPLENDOR", "PULSAR", "CLASSIC 350"])
        elif "gcv" in product or "lcv" in product or "hcv" in product:
            return random.choice(["ACE", "BOLERO MAXI", "DOST"])
        return random.choice(["SWIFT", "ALTO 800", "NEXON", "BALENO"])
        
    elif h == "Fuel Type":
        if "2w" in product or "tw" in product or "two wheeler" in product:
            return "petrol"
        return random.choice(["petrol", "diesel", "CNG"])
        
    elif h == "CPA Status":
        policy_type = str(row_dict.get("Policy Type") or "").lower()
        if "third party" in policy_type or "tp" in policy_type:
            return "NO"
        return random.choice(["YES", "NO"])
        
    elif h == "NCB Status":
        file_type = str(row_dict.get("File Type") or "").lower()
        if "new" in file_type:
            return "NO"
        return random.choice(["YES", "NO"])
        
    elif h == "Partner Type":
        return random.choice(PARTNER_TYPES_POOL)
        
    elif h == "Source":
        return random.choice(SOURCES_POOL)
        
    elif h == "Zone":
        return geo["zone"]
        
    elif h == "RTO":
        return geo["rto"]
        
    elif h == "Vehicle Age From":
        return 1
        
    elif h == "State":
        return random.choice(["MH", "GJ", "AP", "KA", "TN", "DL"])
        
    return val

def build():
    random.seed(42)
    # 10 Real Non-Slab rows
    non_slab_rows = [
        # Row 1: MISC-D TRACTOR
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "MISC-D TRACTOR",
            "Policy Type": "Standalone Own Damage", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Agriculture tractors only without trolly/trailer", "Payin OD": 35.0, "Payin TP": 0.0,
            "Payin Net": 0.0, "Payin Reward": 5.0, "Payin Scheme": 0.0
        },
        # Row 2: PCCV 3W DIESEL, Bajaj
        {
            "LOB": "Motor", "File Type": "Used", "Insurance Company": "Shriram", "Product": "PCCV 3W DIESEL",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "BAJAJ", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 5, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Bajaj Manufacturer Only", "Payin OD": 37.0, "Payin TP": 0.0,
            "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0
        },
        # Row 3: PCCV 3W DIESEL, Other Make
        {
            "LOB": "Motor", "File Type": "Used", "Insurance Company": "Shriram", "Product": "PCCV 3W DIESEL",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 5, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Other than Bajaj Manufacturer", "Payin OD": 32.0, "Payin TP": 0.0,
            "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0
        },
        # Row 4: PCCV 3W DIESEL, Older Bajaj
        {
            "LOB": "Motor", "File Type": "Rollover", "Insurance Company": "Shriram", "Product": "PCCV 3W DIESEL",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "BAJAJ", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 6, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Bajaj Manufacturer Only", "Payin OD": 50.0, "Payin TP": 0.0,
            "Payin Net": 0.0, "Payin Reward": 10.0, "Payin Scheme": 0.0
        },
        # Row 5: PCCV 3W E-Rickshaw
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "PCCV 3W E-Rickshaw",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "electric", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Limit 2000 Watt", "Payin OD": 41.5, "Payin TP": 0.0,
            "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0
        },
        # Row 6: PRIVATE CAR PETROL
        {
            "LOB": "Motor", "File Type": "Renewal", "Insurance Company": "Shriram", "Product": "PRIVATE CAR PETROL",
            "Policy Type": "Standalone Own Damage", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "SA-OD not allowed", "Payin OD": 25.0, "Payin TP": 0.0,
            "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0
        },
        # Row 7: Two Wheeler, Bajaj Bike
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "Two Wheeler",
            "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 5 Year TP", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "<= 180 cc", "Make": "BAJAJ", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP, TS", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Less than or equal to 180CC Vehicle only", "Payin OD": 40.0, "Payin TP": 0.0,
            "Payin Net": 20.0, "Payin Reward": 0.0, "Payin Scheme": 0.0
        },
        # Row 8: Two Wheeler, EV
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "Two Wheeler",
            "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 5 Year TP", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "HONDA", "Model": "ALL", "Fuel Type": "electric", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP, TS", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "EV only", "Payin OD": 45.0, "Payin TP": 0.0,
            "Payin Net": 40.0, "Payin Reward": 0.0, "Payin Scheme": 0.0
        },
        # Row 9: Two Wheeler Petrol Rollover
        {
            "LOB": "Motor", "File Type": "Rollover", "Insurance Company": "Shriram", "Product": "Two Wheeler",
            "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 1 Year TP", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "<= 150 cc", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 6, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "AP ONLY", "Payin OD": 50.0, "Payin TP": 0.0,
            "Payin Net": 17.0, "Payin Reward": 0.0, "Payin Scheme": 0.0
        },
        # Row 10: Short term Mumbai PCCV 4W
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "PCCV 4W",
            "Policy Type": "Standalone Own Damage", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "Upto 6+1", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "MH, GA", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Short Term Policy - Pro rata basis (Minimum Policy Period 90 days)", "Payin OD": 80.0, "Payin TP": 45.0,
            "Payin Net": 0.0, "Payin Reward": 0.0, "Payin Scheme": 0.0
        }
    ]

    # 10 Real Slab rules, containing multiple tiers with non-overlapping continuous ranges.
    # Note: These will use merged cells for columns 1-23.
    slab_rules = [
        # Rule 1: GCCV HCV (State: AP) - 3 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "GCCV HCV",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "HCV Weight Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 12001, "Payin Slab To": 42500, "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0},
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 42501, "Payin Slab To": 50000, "Payin OD": 15.0, "Payin TP": 0.0, "Payin Net": 0.0},
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 50001, "Payin Slab To": 999999, "Payin OD": 12.5, "Payin TP": 0.0, "Payin Net": 0.0}
            ]
        },
        # Rule 2: GCCV LCV (State: AP) - 3 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "GCCV LCV",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "LCV Weight Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2000, "Payin OD": 53.5, "Payin TP": 0.0, "Payin Net": 0.0},
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2001, "Payin Slab To": 2800, "Payin OD": 48.5, "Payin TP": 0.0, "Payin Net": 0.0},
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2801, "Payin Slab To": 3500, "Payin OD": 10.0, "Payin TP": 0.0, "Payin Net": 0.0}
            ]
        },
        # Rule 3: PCCV 4W (State: AP) - 2 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "PCCV 4W",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "PCCV Capacity Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 6, "Payin OD": 27.5, "Payin TP": 0.0, "Payin Net": 0.0},
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 7, "Payin Slab To": 10, "Payin OD": 20.0, "Payin TP": 0.0, "Payin Net": 0.0}
            ]
        },
        # Rule 4: Two Wheeler (State: AP) - 2 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "Two Wheeler",
            "Policy Type": "Comprehensive", "Plan Type": "1 Year OD + 5 Year TP", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "BAJAJ", "Model": "ALL", "Fuel Type": "petrol", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 1, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "2W CC Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 1, "Payin Slab To": 150, "Payin OD": 50.0, "Payin TP": 0.0, "Payin Net": 17.0},
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 151, "Payin Slab To": 180, "Payin OD": 40.0, "Payin TP": 0.0, "Payin Net": 20.0}
            ]
        },
        # Rule 5: GCCV HCV (State: Assam Zone) - 2 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "GCCV HCV",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AS, TR, AR, ML, MZ", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Assam zone HCV Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 12001, "Payin Slab To": 20000, "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0},
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 20001, "Payin Slab To": 42500, "Payin OD": 27.0, "Payin TP": 0.0, "Payin Net": 0.0}
            ]
        },
        # Rule 6: GCCV LCV (State: Assam Zone) - 2 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "GCCV LCV",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 15, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AS, TR, AR, ML, MZ", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Assam zone LCV Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 1, "Payin Slab To": 2500, "Payin OD": 22.5, "Payin TP": 0.0, "Payin Net": 0.0},
                {"Payin Type": "NET", "Premium Type": "OD", "Payin Slab From": 2501, "Payin Slab To": 3500, "Payin OD": 25.0, "Payin TP": 0.0, "Payin Net": 0.0}
            ]
        },
        # Rule 7: GPA (Non-Motor mapped to Motor LOB) - 3 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "GPA",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "GPA Volume Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 1, "Payin Slab To": 500000, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 10.0},
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 500001, "Payin Slab To": 1500000, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 12.5},
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 1500001, "Payin Slab To": 99999999, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 15.0}
            ]
        },
        # Rule 8: Fire & Engineering (Non-Motor mapped to Motor LOB) - 3 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "Fire & Engineering",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Fire Volume Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 1, "Payin Slab To": 100000, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 25.0},
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 100001, "Payin Slab To": 500000, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 27.5},
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 500001, "Payin Slab To": 99999999, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 30.0}
            ]
        },
        # Rule 9: Marine (Non-Motor mapped to Motor LOB) - 2 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "Marine",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Marine Volume Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 1, "Payin Slab To": 200000, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 22.5},
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 200001, "Payin Slab To": 99999999, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 25.0}
            ]
        },
        # Rule 10: Liability/ WC (Non-Motor mapped to Motor LOB) - 2 Tiers
        {
            "LOB": "Motor", "File Type": "New", "Insurance Company": "Shriram", "Product": "Liability/ WC",
            "Policy Type": "Comprehensive", "Plan Type": "ALL", "Sub Product": "NA", "Class": "ALL",
            "Sub Class": "ALL", "Make": "ALL", "Model": "ALL", "Fuel Type": "diesel", "CPA Status": "ALL",
            "NCB Status": "ALL", "Vehicle Age From": 1, "Vehicle Age To": 50, "Partner Type": "ALL partners",
            "Source": "ALL", "Zone": "ALL", "State": "AP", "RTO": "ALL", "Effective Date": "2026-07-13",
            "Remarks": "Liability Volume Slabs",
            "slabs": [
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 1, "Payin Slab To": 500000, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 25.0},
                {"Payin Type": "NET", "Premium Type": "NET", "Payin Slab From": 500001, "Payin Slab To": 99999999, "Payin OD": 0.0, "Payin TP": 0.0, "Payin Net": 27.5}
            ]
        }
    ]

    wb = Workbook()

    # --- NON SLAB SHEET ---
    ws_ns = wb.active
    ws_ns.title = "NON SLAB"
    ws_ns.append(NON_SLAB_HEADERS)
    for cell in ws_ns[1]:
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN, border=THIN_BORDER)

    for r_idx, r in enumerate(non_slab_rows, start=2):
        row_values = []
        for h in NON_SLAB_HEADERS:
            val = r.get(h)
            if val is None:
                if h == "Payout OD" and r.get("Payin OD") is not None:
                    val = r.get("Payin OD") * 0.8
                elif h == "Payout TP" and r.get("Payin TP") is not None:
                    val = r.get("Payin TP") * 0.8
                elif h == "Payout Net" and r.get("Payin Net") is not None:
                    val = r.get("Payin Net") * 0.8
                elif h == "Payout Reward" and r.get("Payin Reward") is not None:
                    val = r.get("Payin Reward") * 0.8
                elif h == "Payout Scheme" and r.get("Payin Scheme") is not None:
                    val = r.get("Payin Scheme") * 0.8
                else:
                    val = "ALL" if h in ["Plan Type", "Make", "Model", "Fuel Type", "CPA Status", "NCB Status", "Source", "Zone", "RTO"] else None
            val = fill_with_realistic_value(h, val, r)
            row_values.append(val)
        ws_ns.append(row_values)
        for col_idx in range(1, len(NON_SLAB_HEADERS) + 1):
            cell = ws_ns.cell(row=r_idx, column=col_idx)
            style_cell(cell, alignment=CENTER_ALIGN, border=THIN_BORDER)

    # --- SLAB SHEET ---
    ws_sl = wb.create_sheet("SLAB")
    ws_sl.append(SLAB_HEADERS)
    for cell in ws_sl[1]:
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN, border=THIN_BORDER)

    current_row = 2
    for r in slab_rules:
        slabs = split_slabs_by_premium_type(r["slabs"])
        num_slabs = len(slabs)
        
        # Write rows first
        for i, slab in enumerate(slabs):
            row_idx = current_row + i
            # Write business columns (LOB to Remarks)
            for col_idx, h in enumerate(SLAB_HEADERS[:23]):
                val = r.get(h)
                if val is None:
                    val = "ALL" if h in ["Plan Type", "Make", "Model", "Fuel Type", "CPA Status", "NCB Status", "Source", "Zone", "RTO"] else None
                val = fill_with_realistic_value(h, val, r)
                cell = ws_sl.cell(row=row_idx, column=col_idx + 1, value=val)
                style_cell(cell, alignment=CENTER_ALIGN, border=THIN_BORDER)
                
            # Write slab columns
            for col_idx, h in enumerate(SLAB_HEADERS[23:]):
                val = slab.get(h)
                if val is None:
                    if h == "Payout OD" and slab.get("Payin OD") is not None:
                        val = slab.get("Payin OD") * 0.8
                    elif h == "Payout TP" and slab.get("Payin TP") is not None:
                        val = slab.get("Payin TP") * 0.8
                    elif h == "Payout Net" and slab.get("Payin Net") is not None:
                        val = slab.get("Payin Net") * 0.8
                cell = ws_sl.cell(row=row_idx, column=24 + col_idx, value=val)
                style_cell(cell, alignment=CENTER_ALIGN, border=THIN_BORDER)
                
        # Apply Merges for business columns
        if num_slabs > 1:
            for col_idx in range(1, 24):
                ws_sl.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row + num_slabs - 1, end_column=col_idx)
                # Re-apply alignment/border styling to merged range cells to keep design premium
                for row_offset in range(num_slabs):
                    cell = ws_sl.cell(row=current_row + row_offset, column=col_idx)
                    style_cell(cell, alignment=CENTER_ALIGN, border=THIN_BORDER)
                    
        current_row += num_slabs

    autosize(ws_ns)
    autosize(ws_sl)

    try:
        wb.save(OUT_PATH)
        print(f"Successfully generated Shriram Excel file at: {OUT_PATH}")
    except PermissionError:
        print(f"ERROR: Permission denied when writing to {OUT_PATH}. Please close this file in Microsoft Excel and run again!")

if __name__ == "__main__":
    build()
