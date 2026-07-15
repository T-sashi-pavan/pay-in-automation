"""
Covers excel_export.py's build_export_workbook — the real .xlsx download that
replaced the old client-side CSV/JSON export. Uses stub rule/slab objects
(no DB needed: product/state are None here, so master_data_service's lookups
short-circuit without touching the db session passed in).
"""
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from backend.app.services.excel_export import build_export_workbook


def _make_slab(**overrides):
    defaults = dict(id=1, payin_type="PERCENTAGE", premium_type="OD", slab_from=0, slab_to=40, payin_od=30.0, payin_tp=None, payin_net=None)
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def _make_rule(**overrides):
    defaults = dict(
        id=1, upload_id=1, sheet_name="Sheet1",
        lob=None, file_type=None, insurance_company="Shriram", product=None,
        policy_type="Comprehensive", plan_type=None, sub_product=None, class_=None,
        sub_class="GVW 3501-7500", make=None, model=None, fuel_type=None, body_type=None,
        vehicle_age_from=None, vehicle_age_to=10, cpa_status=None, ncb_status="YES",
        partner_type=None, state=None, zone=None, source=None, rto=None,
        effective_date=None, remarks=None,
        commission_type="NON_SLAB", slab_configuration=False,
        payin_od=80.0, payout_od=None, payin_tp=None, payout_tp=None,
        payin_net=None, payout_net=None, payin_reward=None, payout_reward=None,
        payin_scheme=None, payout_scheme=None,
        validation_status="VALID", warnings=[], raw_json={}, slabs=[],
    )
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def test_export_workbook_has_two_sheets_with_correct_row_split():
    non_slab_rule = _make_rule(id=1, commission_type="NON_SLAB")
    slab_rule = _make_rule(id=2, commission_type="SLAB", slab_configuration=True, slabs=[
        _make_slab(slab_from=0, slab_to=40),
        _make_slab(slab_from=41, slab_to=70, premium_type="TP", payin_od=None, payin_tp=25.0),
    ])

    buffer = build_export_workbook([non_slab_rule, slab_rule], db=None)
    wb = load_workbook(BytesIO(buffer.read()))

    assert wb.sheetnames == ["Non-Slab", "Slab"]
    assert wb["Non-Slab"].max_row == 2  # header + 1 non-slab rule

    # New vertical layout: 1 header row + 1 row per slab tier of each rule.
    # The slab_rule has 2 tiers → 1 header + 2 tier rows = max_row 3.
    assert wb["Slab"].max_row == 3      # 1 header + 2 tier rows (one row per slab tier)

    # Verify both tier rows carry slab data (slab-detail cols start at n_biz+1 = col 25)
    ws_sl = wb["Slab"]
    # Row 2 = tier 1 (slab_from=0, premium_type=OD)
    # Row 3 = tier 2 (slab_from=41, premium_type=TP)
    # Payin Slab From is the 3rd slab col → col 24+3 = 27
    from backend.app.services.excel_export import BUSINESS_COLUMNS
    slab_from_col = len(BUSINESS_COLUMNS) + 3  # Pay-In Type, Premium Type, then Payin Slab From
    tier1_from = ws_sl.cell(row=2, column=slab_from_col).value
    tier2_from = ws_sl.cell(row=3, column=slab_from_col).value
    assert tier1_from == 0    # first tier (lowest slab_from after sort)
    assert tier2_from == 41   # second tier


def test_export_workbook_flags_defaulted_cells_in_purple():
    rule = _make_rule(id=1, lob=None)  # lob is null on the stub -> should be defaulted/purple
    buffer = build_export_workbook([rule], db=None)
    wb = load_workbook(BytesIO(buffer.read()))

    ws = wb["Non-Slab"]
    lob_cell = ws.cell(row=2, column=1)  # LOB is the first business column
    assert lob_cell.value == "Motor"  # default_or's fallback value
    assert lob_cell.font.color.rgb in ("006D28D9", "FF6D28D9")  # the purple "this is a default" color

    policy_type_cell = ws.cell(row=2, column=5)  # Policy Type — genuinely populated on the stub
    assert policy_type_cell.value == "Comprehensive"
    assert policy_type_cell.font.color is None or policy_type_cell.font.color.rgb not in ("006D28D9", "FF6D28D9")
