import openpyxl
import os

FILES_TO_TRIM = [
    ("Shriram JUL'26 BROKER GRID.xlsx", "unstructured_sample_shriram.xlsx", 30),
    ("Chola 21st MAY Broking Grid May 25.xlsx", "unstructured_sample_chola.xlsx", 30),
    ("Go digit SQUARE_May25.xlsx", "unstructured_sample_digit.xlsx", 30),
    ("Tata Gird GRID JULY---.xlsx", "unstructured_sample_tata.xlsx", 30),
]

def trim_files():
    for src, dst, max_keep_rows in FILES_TO_TRIM:
        src_path = os.path.join(os.getcwd(), src)
        dst_path = os.path.join(os.getcwd(), dst)
        if not os.path.exists(src_path):
            print(f"Skipping {src} (not found)")
            continue
            
        print(f"Trimming {src} -> {dst}")
        wb = openpyxl.load_workbook(src_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            # Print row count before
            orig_rows = ws.max_row
            if orig_rows > max_keep_rows:
                ws.delete_rows(max_keep_rows + 1, orig_rows - max_keep_rows)
                print(f"  Sheet '{sheet}': deleted {orig_rows - max_keep_rows} rows (kept top {max_keep_rows})")
            else:
                print(f"  Sheet '{sheet}': kept all {orig_rows} rows")
        
        try:
            wb.save(dst_path)
            print(f"Successfully saved unstructured sample to: {dst_path}")
        except PermissionError:
            print(f"ERROR: Permission denied when saving to {dst_path}. Please close this file if open!")

if __name__ == '__main__':
    trim_files()
